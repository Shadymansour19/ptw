"""Generate printable reports for PTWs and ICs.

Builds ReportLab PDF permit/certificate reports (each with an embedded QR-code
payload and a printed watermark), the Method of Statement and specific risk
assessment PDF supplements, and an openpyxl Excel export of a PTW list — opening
each generated file in the system's default viewer once built.
"""

import html as _html
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import *
from reportlab.lib import colors
from reportlab.lib.styles import *
from reportlab.lib.enums import *
from reportlab.platypus import *
from reportlab.lib.units import mm, inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import qrcode
from pypdf import PdfWriter, PaperSize, Transformation
from pypdf.generic import RectangleObject
from PIL import Image as PILImage, ImageDraw as PILImageDraw
import webbrowser
from datetime import datetime
from GlobalData import globalData
from models.PTW import PTW, RiskAssessment
from models.Isolation import IC
import io
import tempfile
import platform
import os
import subprocess
from network.clientRequests import ClientRequests
from helper.utils import resource_path
from reports.ArabicText import pdfMarkup, isRtlBase, REGULAR_FONT_NAME, BOLD_FONT_NAME


class ReportGenerator:
    """Namespace of report-building routines invoked directly on the class: PDF
    PTW/IC reports, MOS and risk-assessment PDF supplements, and an Excel PTW
    export."""

    def _registerArabicFonts():
        """Register the bundled Noto Naskh Arabic fonts with ReportLab, once - every
        report-building method calls this before building any Paragraph, since none
        of ReportLab's built-in fonts (Helvetica et al.) carry Arabic glyphs at all
        (see reports/ArabicText.py, which tags Arabic runs with these font names)."""
        try:
            pdfmetrics.getFont(REGULAR_FONT_NAME)
        except KeyError:
            pdfmetrics.registerFont(TTFont(REGULAR_FONT_NAME, resource_path('fonts/NotoNaskhArabic/NotoNaskhArabic-Regular.ttf')))
            pdfmetrics.registerFont(TTFont(BOLD_FONT_NAME, resource_path('fonts/NotoNaskhArabic/NotoNaskhArabic-Bold.ttf')))

    def arabicParagraph(text, style, forceAlignment=True):
        """Build a `Paragraph` from `text`, the Arabic-aware way: bidi-reorder/reshape
        and Arabic-font-tag any Arabic-script runs (see reports/ArabicText.py), and -
        if `forceAlignment` - right-align the whole paragraph when its own base
        direction is Arabic, regardless of what language the rest of the report (or
        the app) is in. A drop-in replacement for `Paragraph(html.escape(text), style)`
        wherever `text` is a plain (non-markup) string that might be Arabic, English,
        or a mix of both - every printed field is free text a user typed, so any of
        them can be. Pass `forceAlignment=False` for a slot whose own centered/other
        alignment should win regardless of content (e.g. a narrow signature-block
        cell) - the Arabic shaping/font-tagging still applies either way."""
        text = text if text is not None else ''
        if forceAlignment and isRtlBase(text):
            style = ParagraphStyle(f'{style.name}-RTL', parent=style, alignment=TA_RIGHT)
        return Paragraph(pdfMarkup(text), style)

    def _qrWithLogoFromRows(basicInfo: list, filePrefix: str):
        """Build a QR code encoding basicInfo, with the app logo composited over
        its center in a white circle, and save it to a temporary PNG.

        Args:
            basicInfo: list of [label, value] pairs, joined into 'label: value'
                lines to form the QR payload.
            filePrefix: prefix used for the temp file name.

        Returns:
            str: path to the generated PNG file.
        """
        data = '\n'.join([': '.join(row) for row in basicInfo])
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_Q, box_size=20, border=2)
        qr.add_data(data)
        qr.make(fit=True)
        qrImg = qr.make_image(fill_color="black", back_color="white").convert("RGB")
        qr_w, qr_h = qrImg.size

        logo_size = qr_w // 3
        padding = logo_size // 32
        logo = PILImage.open(resource_path("assets/sh-logo-bw.png")).convert("RGB").resize((logo_size, logo_size), PILImage.LANCZOS)
        logo_mask = PILImage.new("L", (logo_size, logo_size), 0)
        PILImageDraw.Draw(logo_mask).ellipse((0, 0, logo_size - 1, logo_size - 1), fill=255)
        total_size = logo_size + 2 * padding
        bg = PILImage.new("RGB", (total_size, total_size), "white")
        bg_mask = PILImage.new("L", (total_size, total_size), 0)
        PILImageDraw.Draw(bg_mask).ellipse((0, 0, total_size - 1, total_size - 1), fill=255)
        bg.paste(logo, (padding, padding), logo_mask)
        qrImg.paste(bg, ((qr_w - total_size) // 2, (qr_h - total_size) // 2), bg_mask)
        with tempfile.NamedTemporaryFile(delete=False, prefix=f'qr-{filePrefix}-', suffix='.png') as qrFile:
            qrImg.save(qrFile.name)
            return qrFile.name

    def _makeQrWithLogo(ptw: PTW):
        """Build the logo QR code encoding a PTW's basic info (id, type, status via
        `runningStatusDisplay()`, department, requestor, PA, location, equipment,
        description).

        Returns:
            str: path to the generated PNG file.
        """
        basicInfo = [
            ['PTW#', str(ptw.id)],
            ['Type', str(ptw.type)],
            ['Status', ptw.runningStatusDisplay()],
            ['Department', str(ptw.department)],
            ['Requestor', str(globalData.allUsers[ptw.requestor].getName()) if ptw.requestor in globalData.allUsers else 'None'],
            ['PA', str(globalData.allUsers[ptw.getPerforming()].getName()) if ptw.getPerforming() in globalData.allUsers else 'None'],
            ['Location', str(ptw.location)],
            ['Equipment', str(ptw.equipment)],
            ['Description', str(ptw.description)],
        ]
        return ReportGenerator._qrWithLogoFromRows(basicInfo, str(ptw.id))

    def _makeQrWithLogoIC(ic: IC):
        """Build the logo QR code encoding an IC's basic info (id, type, status,
        requestor/execution department, requestor, location, equipment, reason).

        Returns:
            str: path to the generated PNG file.
        """
        basicInfo = [
            ['IC#', str(ic.id)],
            ['Type', str(ic.type)],
            ['Status', str(ic.getStatus())],
            ['Requestor Dept', str(ic.requestor_department)],
            ['Execution Dept', str(ic.execution_department)],
            ['Requestor', str(globalData.allUsers[ic.requestor].getName()) if ic.requestor in globalData.allUsers else 'None'],
            ['Location', str(ic.location)],
            ['Equipment', str(ic.equipment)],
            ['Reason', str(ic.reason)],
        ]
        return ReportGenerator._qrWithLogoFromRows(basicInfo, f'ic-{ic.id}')


    def ptwReport(loggedUser, ptw: PTW):
        """Build and open the PTW permit PDF, then open its related documents.

        Lays out a landscape-A4 PDF with the logo QR code (see `_makeQrWithLogo`)
        printed in the left margin of every page and a diagonal "Printed @
        <timestamp>" watermark: a Summary table of basic PTW fields (status shown
        via `ptw.runningStatusDisplay()`), an Additional Info table listing
        tools/hazards/controls/MIWI/attachments (attachment names are fetched live
        from the server rather than trusted from `ptw.attachs`), the Method of
        Statement as a bulleted list when present, and signature tables for the
        approval chain (`ptw.requiredApprovers()`) and the PA/IA running
        confirmations. After writing and opening the PDF, it also fetches and opens
        the PTW's MIWI document, the PTW-specific risk assessment report, and every
        uploaded attachment, plus the PTW's bundled required documents
        (`ptw.requiredDocsToPrint()`, resolved from reports/docs/) merged into a
        single PDF.

        Args:
            loggedUser: user requesting the report, used to authorize the server
                fetches for risk assessment, attachments, and MIWI.
            ptw: the PTW being reported on.

        Returns:
            An error string if the MIWI document couldn't be fetched, else None.
        """
        ReportGenerator._registerArabicFonts()

        QR_CODE_WIDTH = 60*mm
        LOGO_IMG_WIDTH = 40*mm
        TABLE_LABELS_WIDTH = 50*mm
        MARGIN = 0.6 * inch

        buffer = io.BytesIO()
        timestamp = datetime.now().strftime('%d-%m-%Y - %H:%M:%S')
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=MARGIN+QR_CODE_WIDTH, rightMargin=MARGIN, topMargin=MARGIN*0.7, bottomMargin=MARGIN*0.7)

        pageWidth, pageHeight = landscape(A4)
        dataTableWidth = pageWidth - 2*MARGIN - QR_CODE_WIDTH

        styles = getSampleStyleSheet()
        
        styles['Title'].fontSize = 24
        styles['Title'].leading = 26
        styles["Title"].alignment = TA_LEFT

        styles['Normal'].fontSize = 16
        styles['Normal'].leading = 18
        styles['Normal'].fontName = 'Helvetica'

        styles['Heading3'].fontSize = 16
        styles['Heading3'].leading = 18
        styles['Heading3'].fontName = 'Helvetica-Bold'

        def listTo2ColsBullets(data: list, style):
            """Split data into two roughly equal halves and lay them out as a
            two-column table of bulleted, Arabic-aware paragraphs (see
            ReportGenerator.arabicParagraph)."""
            mid = (len(data) + 1) // 2
            col1Data = data[:mid]
            col2Data = data[mid:]
            tableData = [
                [ReportGenerator.arabicParagraph('• ' + col1Data[i], style), (ReportGenerator.arabicParagraph('• ' + col2Data[i], style) if i < len(col2Data) else None)]
                for i in range(len(col1Data))
            ]
            return Table(tableData)

        def listToBullets(data: list, style):
            """Render a list of strings as a bulleted ListFlowable of Arabic-aware
            paragraphs (see ReportGenerator.arabicParagraph), or None if data is empty."""
            if not data:
                return None

            bullets = [ListItem(ReportGenerator.arabicParagraph(element, style), bulletType='bullet', bulletFontSize=style.fontSize) for element in data]
            return ListFlowable(
                bullets, 
                bulletType='bullet', 
                leftIndent=0.3*inch, 
                bulletIndent=0.1*inch, 
                spaceAfter=0.1*inch, 
            )

        basicInfo = [
            ['PTW#', str(ptw.id)], 
            ['Type', str(ptw.type)], 
            ['Print Time', timestamp], 
            ['Status', ptw.runningStatusDisplay()], 
            ['Request Date', str(ptw.request_date)], 
            ['Department', str(ptw.department)], 
            ['Requestor', str(globalData.allUsers[ptw.requestor].getName()) if ptw.requestor in globalData.allUsers else 'None'],
            ['PA', str(globalData.allUsers[ptw.getPerforming()].getName()) if ptw.getPerforming() in globalData.allUsers else 'None'],
            ['Location', str(ptw.location)],
            ['Equipment', str(ptw.equipment)],
            ['Area Class', str(ptw.area_class)],
            ['Description', str(ptw.description)], 
        ]

        qrPath = ReportGenerator._makeQrWithLogo(ptw)

        elements = []

        def insertTable(title, tableData: list[list]):
            """Append a titled two-column label/value table (with a shaded label
            column and full grid) to `elements`, followed by a page break. No-op
            if tableData is empty."""
            if not tableData:
                return
            
            nonlocal elements
            table = Table(tableData, colWidths=[TABLE_LABELS_WIDTH, dataTableWidth - TABLE_LABELS_WIDTH], splitByRow=1)
            table.setStyle(TableStyle([
                # header column (left)
                ('BACKGROUND', (0, 0), (0, -1), colors.Color(0, 0, 0, 0.2)),

                # Alignment
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'), 
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), 

                # padding
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),

                # grid (optional)
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.extend([title, Spacer(1, 0.2 * inch), table, PageBreak()])

        insertTable(Paragraph('Summery:', styles["Title"]), [
            [Paragraph(row[0], styles['Heading3']), ReportGenerator.arabicParagraph(row[1], styles['Normal'])]
            for row in basicInfo
        ])

        err, ptwSpecificRisk = ClientRequests.getPTWSpecificRiskAssessment(loggedUser, ptw.id)
        if err:
            ptwSpecificRisk = None

        # Attachment filenames are never persisted on the PTW row itself, only the files
        # on disk (ptw-{id}-attachments/) are authoritative — fetch the live listing rather
        # than trusting ptw.attachs (which reflects the local, not-yet-uploaded staging list).
        err, attachNames = ClientRequests.getPtwAttachmentNames(loggedUser, ptw.id)
        if err:
            attachNames = []

        tableAdditionalInfoData = [
            [Paragraph('Tools', styles['Heading3']), listToBullets(ptw.tools, styles['Normal'])],
            [Paragraph('Hazards', styles['Heading3']), listToBullets(ptw.hazards, styles['Normal'])],
            [Paragraph('Controls', styles['Heading3']), listToBullets(ptw.controls, styles['Normal'])],
        ]
        if ptw.miwi:
            tableAdditionalInfoData.append([Paragraph('MIWI', styles['Heading3']), listToBullets([ptw.miwi], styles['Normal'])])
        # elif ptw.mos:
        #     tableAdditionalInfoData.append([Paragraph('MOS', styles['Heading3']), listToBullets(ptw.mos.split('\n'), styles['Normal'])])
        if attachNames:
            tableAdditionalInfoData.append([Paragraph('Attachments', styles['Heading3']), listToBullets(attachNames, styles['Normal'])])
        # if ptw.isolations:
        #     tableAdditionalInfoData.append([Paragraph('Isolations', styles['Heading3']), listToBullets([str(isolation) for isolation in ptw.isolations], styles['Normal'])])

        insertTable(Paragraph('Additional Info:', styles["Title"]), tableAdditionalInfoData)


        if ptw.mos:
            mosSteps = ptw.mos.split('\n')
            elements.extend([Paragraph('Method Of Statement:', styles["Title"]), Spacer(1, 0.2 * inch)])
            elements.append(listToBullets(mosSteps, styles['Normal']))
            elements.append(PageBreak())
        

        try:
            pdfmetrics.registerFont(TTFont('Satisfy', resource_path('fonts/Satisfy/Satisfy-Regular.ttf')))
            sig_font = 'Satisfy'
        except Exception:
            sig_font = 'Helvetica-Oblique'

        label_style = ParagraphStyle('SigLabel', parent=styles['Heading3'], fontSize=13, leading=14, alignment=TA_CENTER)
        sig_style   = ParagraphStyle('Signature', parent=styles['Normal'],  fontSize=16, leading=17, alignment=TA_CENTER, fontName=sig_font)
        date_style  = ParagraphStyle('SigDate',   parent=styles['Normal'],  fontSize=12,  leading=14, alignment=TA_CENTER)

        def sigTables(columns: list, chunk_size=3):
            """columns: list of (label, sig, ts) or None for empty padding slots."""
            chunks = [columns[i:i+chunk_size] for i in range(0, len(columns), chunk_size)]
            result = []
            for chunk in chunks:
                col_w = dataTableWidth / len(chunk)
                header_row, name_row, date_row = [], [], []
                style_cmds = [
                    ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING',    (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]
                for i, col in enumerate(chunk):
                    if col is not None:
                        label, sig, ts = col
                        style_cmds.append(('LINEBELOW', (i, 0), (i, 0), 0.5, colors.black))
                        style_cmds.append(('LINEBELOW', (i, 1), (i, 1), 0.5, colors.black))
                    else:
                        label, sig, ts = '', '', ''
                    header_row.append(Paragraph(label or '\u00a0',        label_style))
                    name_row.append(  ReportGenerator.arabicParagraph(sig or '\u00a0', sig_style, forceAlignment=False))
                    date_row.append(  Paragraph(ts    or '\u00a0',        date_style))
                table = Table(
                    [header_row, name_row, date_row],
                    colWidths=[col_w] * len(chunk),
                )
                table.setStyle(TableStyle(style_cmds))
                result.append(table)
                result.append(Spacer(1, 0.09 * inch))
            return result

        def approvalColumns():
            """Build one signature column per required approver stage/slot, padded
            to 7 with blanks: (approver label, approver's name if approved, else
            '', approval timestamp if approved, else '')."""
            def lastApprovalFor(approver):
                """Return the most recent approval on ptw matching approver, or None."""
                match = None
                for approval in ptw.approvals:
                    if approver.matchesUser(globalData.allUsers.get(approval.username)):
                        match = approval
                return match

            cols = []
            for stage in ptw.requiredApprovers():
                for approver in stage:
                    approval = lastApprovalFor(approver)
                    if approval and approval.action == PTW.ApprovalActions.APPROVED:
                        name = globalData.allUsers[approval.username].getName()
                        ts   = approval.timestamp or ''
                    else:
                        name, ts = '', ''
                    cols.append((str(approver), name, ts))
            while len(cols) < 7:
                cols.append(None)
            return cols

        def runConfirmationColumns():
            """Build the PA/IA signature columns for the currently open run cycle:
            (role label, signer's name, signature timestamp) for each of PA and IA."""
            performing = ptw.getPerforming()
            issuing = ptw.getIssuing()
            pa = globalData.allUsers[performing].getName() if performing in globalData.allUsers else str(performing or '')
            ia = globalData.allUsers[issuing].getName()    if issuing    in globalData.allUsers else str(issuing    or '')
            return [
                ('PA', pa, ptw.getPerformingTimestamp() or ''),
                ('IA', ia, ptw.getIssuingTimestamp()    or ''),
            ]

        elements.extend([
            Paragraph('Approvals:', styles["Title"]),
            Spacer(1, 0.07 * inch),
            *sigTables(approvalColumns()),
            Spacer(1, 0.1 * inch),
            Paragraph('Running Confirmations:', styles["Title"]),
            Spacer(1, 0.07 * inch),
            *sigTables(runConfirmationColumns()),
            PageBreak(),
        ])

        # insertTable(Paragraph('Approval Cycle:', styles["Title"]), [
        #     [Paragraph(approval.action, styles['Heading3']), Paragraph(str(approval), styles['Normal'])]
        #     for approval in ptw.approvals
        # ])
        

        class NumberedCanvas(canvas.Canvas):
            """ReportLab canvas that defers page-number drawing until save(), so
            each page can display "Page X of Y" with the final total page count."""

            def __init__(self, *args, **kwargs):
                """Initialize the underlying canvas and the buffer of saved page states."""
                canvas.Canvas.__init__(self, *args, **kwargs)
                self._saved_page_states = []

            def showPage(self):
                """Stash the current page's drawing state instead of finalizing the
                page immediately, so the page-number footer can be added later."""
                self._saved_page_states.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                """Replay each saved page state, draw its "Page X of Y" footer, then
                finalize all pages and save the document."""
                num_pages = len(self._saved_page_states)
                for state in self._saved_page_states:
                    self.__dict__.update(state)
                    self.setFont('Helvetica', 14)
                    self.setFillColorRGB(0, 0, 0, 1)
                    self.drawCentredString(0.7 * MARGIN + QR_CODE_WIDTH / 2, MARGIN, f'Page {self._pageNumber} of {num_pages}')
                    canvas.Canvas.showPage(self)
                canvas.Canvas.save(self)

        def pageHeaderAndWatermark(canvas: canvas.Canvas, doc):
            """Draw the two company logos, a diagonal "Printed @ <timestamp>"
            watermark, and the logo QR code into the page's left margin. Used as
            the ReportLab onFirstPage/onLaterPages callback."""
            canvas.saveState()

            canvas.drawImage(resource_path("assets/rashpetco-logo.png"), 0.7 * MARGIN + (QR_CODE_WIDTH - LOGO_IMG_WIDTH) / 2, (pageHeight - LOGO_IMG_WIDTH) / 2.0 + pageHeight / 3.5, LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, mask='auto')
            canvas.drawImage(resource_path("assets/burullus-logo.png"),  0.7 * MARGIN + (QR_CODE_WIDTH - LOGO_IMG_WIDTH) / 2, (pageHeight - LOGO_IMG_WIDTH) / 2.0 - pageHeight / 3.5, LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, mask='auto')

            canvas.setFont('Helvetica-Bold', 50)
            canvas.setFillColorRGB(0, 0, 0, 0.2)
            canvas.translate(pageWidth / 2.0, pageHeight / 2.0)
            canvas.rotate(35)
            canvas.drawCentredString(0, 0, f'Printed @ {timestamp}')

            canvas.restoreState()
            canvas.drawImage(qrPath, 0.7 * MARGIN, (pageHeight - QR_CODE_WIDTH) / 2.0, QR_CODE_WIDTH, QR_CODE_WIDTH)

        # toolsTitle = Paragraph('Tools:', styles["Title"])
        # toolsBullets = [ListItem(Paragraph(tool, styles['Normal']), bulletType='bullet', bulletFontSize=styles['Normal'].fontSize) for tool in ptw.tools]
        # toolsBulletsList = ListFlowable(
        #     toolsBullets, 
        #     bulletType='bullet', 
        #     leftIndent=0.3*inch, 
        #     bulletIndent=0.1*inch, 
        #     spaceAfter=0.1*inch, 
        # )
        
        # hazardsTitle = Paragraph('Hazards:', styles["Title"])
        # hazardsBullets = [ListItem(Paragraph(tool, styles['Normal']), bulletType='bullet', bulletFontSize=styles['Normal'].fontSize) for tool in ptw.hazards]
        # hazardsBulletsList = ListFlowable(
        #     hazardsBullets, 
        #     bulletType='bullet', 
        #     leftIndent=0.3*inch, 
        #     bulletIndent=0.1*inch, 
        #     spaceAfter=0.1*inch, 
        # )

        # controlsTitle = Paragraph('Controls:', styles["Title"])
        # controlsBullets = [ListItem(Paragraph(tool, styles['Normal']), bulletType='bullet', bulletFontSize=styles['Normal'].fontSize) for tool in ptw.controls]
        # controlsBulletsList = ListFlowable(
        #     controlsBullets, 
        #     bulletType='bullet', 
        #     leftIndent=0.3*inch, 
        #     bulletIndent=0.1*inch, 
        #     spaceAfter=0.1*inch, 
        # )

        # risksTitle = Paragraph('Risks:', styles["Title"])
        # risksBullets = [ListItem(Paragraph(tool, styles['Normal']), bulletType='bullet', bulletFontSize=styles['Normal'].fontSize) for tool in ptw.risks]
        # risksBulletsList = ListFlowable(
        #     risksBullets, 
        #     bulletType='bullet', 
        #     leftIndent=0.3*inch, 
        #     bulletIndent=0.1*inch, 
        #     spaceAfter=0.1*inch, 
        # )

        doc.build(elements, onFirstPage=pageHeaderAndWatermark, onLaterPages=pageHeaderAndWatermark, canvasmaker=NumberedCanvas)
        buffer.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, prefix=f'ptw-{ptw.id}-', suffix='.pdf') as ptwPdfFile:
            ptwPdfFile.write(buffer.read())
            ptwPdfFile.flush()
            ReportGenerator.openPDF(ptwPdfFile.name)
        # ReportGenerator.MOSReport(ptw.mos, ptw.id, ptw.description)
        if ptw.miwi:
            err, filepath = ClientRequests.getMIWI(loggedUser, ptw.miwi, department=ptw.department)
            if err:
                return err
            else:
                ReportGenerator.openPDF(filepath)
        ReportGenerator.riskAssessmentReport(ptwSpecificRisk)

        err, attachs = ClientRequests.getPtwAttachmentNames(loggedUser, ptw.id)
        if err:
            attachs = []

        for attachment in attachs:
            err, filepath = ClientRequests.getPtwAttachment(loggedUser, ptw.id, attachment)
            if err:
                print(f"Failed to fetch attachment '{attachment}': {err}")
            else:
                ReportGenerator.openPDF(filepath)

        requiredDocs = PdfWriter()
        for docKey in ptw.requiredDocsToPrint():
            docPath = resource_path(os.path.join('reports', 'docs', f'{docKey}.pdf'))
            if os.path.isfile(docPath):
                requiredDocs.append(docPath)
            else:
                print(f"Missing bundled required doc '{docKey}': {docPath}")
        if len(requiredDocs.pages):
            # Normalize every page to one uniform size and orientation - portrait A4.
            # Landscape pages (the audit and gas-test forms) are first rotated with
            # /Rotate 270, putting their content's top at the portrait page's left
            # edge (the binding convention where turning the sheet clockwise makes it
            # readable), then each page is uniformly scaled to fit A4 and centered on
            # an exact-A4 page: the A3 toolbox form shrinks 1:1 (same aspect ratio),
            # while the letter-size audit form keeps its proportions and just gains
            # small margins.
            A4_WIDTH, A4_HEIGHT = float(PaperSize.A4.width), float(PaperSize.A4.height)
            for page in requiredDocs.pages:
                displayedWidth, displayedHeight = float(page.mediabox.width), float(page.mediabox.height)
                if page.rotation % 180:
                    displayedWidth, displayedHeight = displayedHeight, displayedWidth
                if displayedWidth > displayedHeight:
                    page.rotate(270)
                if page.rotation:
                    page.transfer_rotation_to_content()
                pageWidth, pageHeight = float(page.mediabox.width), float(page.mediabox.height)
                scale = min(A4_WIDTH / pageWidth, A4_HEIGHT / pageHeight)
                page.scale_by(scale)
                page.add_transformation(Transformation().translate(
                    (A4_WIDTH - pageWidth * scale) / 2,
                    (A4_HEIGHT - pageHeight * scale) / 2,
                ))
                page.mediabox = RectangleObject([0, 0, A4_WIDTH, A4_HEIGHT])
                page.cropbox = RectangleObject([0, 0, A4_WIDTH, A4_HEIGHT])
            with tempfile.NamedTemporaryFile(delete=False, prefix=f'ptw-{ptw.id}-required-docs-', suffix='.pdf') as requiredDocsFile:
                requiredDocs.write(requiredDocsFile)
            ReportGenerator.openPDF(requiredDocsFile.name)

        return None


    def icReport(loggedUser, ic: IC):
        """Build and open the IC (Isolation Certificate) PDF report.

        Lays out a landscape-A4 PDF with the logo QR code (see `_makeQrWithLogoIC`)
        and a diagonal "Printed @ <timestamp>" watermark on every page: a Summary
        table of basic IC fields, an Isolation Items table (tag/description/state/
        lock number/lock box number), a bulleted list of linked PTW ids, signature
        tables for the approval chain (`ic.requiredApprovers()`), and signature
        tables for each of the isolate/sanction-for-test/re-isolate/de-isolate
        request-confirm-carry-out sequences.

        Args:
            loggedUser: unused by this report but kept for a consistent call
                signature with `ptwReport`.
            ic: the IC being reported on.

        Returns:
            None.
        """
        ReportGenerator._registerArabicFonts()

        QR_CODE_WIDTH = 60*mm
        LOGO_IMG_WIDTH = 40*mm
        TABLE_LABELS_WIDTH = 50*mm
        MARGIN = 0.6 * inch

        buffer = io.BytesIO()
        timestamp = datetime.now().strftime('%d-%m-%Y - %H:%M:%S')
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=MARGIN+QR_CODE_WIDTH, rightMargin=MARGIN, topMargin=MARGIN*0.7, bottomMargin=MARGIN*0.7)

        pageWidth, pageHeight = landscape(A4)
        dataTableWidth = pageWidth - 2*MARGIN - QR_CODE_WIDTH

        styles = getSampleStyleSheet()

        styles['Title'].fontSize = 24
        styles['Title'].leading = 26
        styles["Title"].alignment = TA_LEFT

        styles['Normal'].fontSize = 16
        styles['Normal'].leading = 18
        styles['Normal'].fontName = 'Helvetica'

        styles['Heading3'].fontSize = 16
        styles['Heading3'].leading = 18
        styles['Heading3'].fontName = 'Helvetica-Bold'

        styles.add(ParagraphStyle('NormalCenter', parent=styles['Normal'], alignment=TA_CENTER))
        styles.add(ParagraphStyle('Heading3Center', parent=styles['Heading3'], alignment=TA_CENTER))

        def nameFor(username):
            """Return the display name for username, or '' if username is falsy,
            or the raw username if it's not a known user."""
            if not username:
                return ''
            return globalData.allUsers[username].getName() if username in globalData.allUsers else str(username)

        def listToBullets(data: list, style):
            """Render a list of strings as a bulleted ListFlowable of Arabic-aware
            paragraphs (see ReportGenerator.arabicParagraph), or None if data is empty."""
            if not data:
                return None

            bullets = [ListItem(ReportGenerator.arabicParagraph(element, style), bulletType='bullet', bulletFontSize=style.fontSize) for element in data]
            return ListFlowable(
                bullets,
                bulletType='bullet',
                leftIndent=0.3*inch,
                bulletIndent=0.1*inch,
                spaceAfter=0.1*inch,
            )

        basicInfo = [
            ['IC#', str(ic.id)],
            ['Type', str(ic.type)],
            ['Print Time', timestamp],
            ['Status', str(ic.getStatus())],
            ['Request Date', str(ic.requestor_timestamp)],
            ['Requestor Dept', str(ic.requestor_department)],
            ['Execution Dept', str(ic.execution_department)],
            ['Requestor', nameFor(ic.requestor)],
            ['Location', str(ic.location)],
            ['Equipment', str(ic.equipment)],
            ['Reason', str(ic.reason)],
            ['Long Term', (f'Yes - {ic.long_term_reason}' if ic.long_term_reason else 'Yes') if ic.long_term else 'No'],
        ]

        qrPath = ReportGenerator._makeQrWithLogoIC(ic)

        elements = []

        def insertTable(title, tableData: list[list]):
            """Append a titled two-column label/value table (with a shaded label
            column and full grid) to `elements`, followed by a page break. No-op
            if tableData is empty."""
            if not tableData:
                return

            nonlocal elements
            table = Table(tableData, colWidths=[TABLE_LABELS_WIDTH, dataTableWidth - TABLE_LABELS_WIDTH], splitByRow=1)
            table.setStyle(TableStyle([
                # header column (left)
                ('BACKGROUND', (0, 0), (0, -1), colors.Color(0, 0, 0, 0.2)),

                # Alignment
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

                # padding
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),

                # grid (optional)
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.extend([title, Spacer(1, 0.2 * inch), table, PageBreak()])

        insertTable(Paragraph('Summery:', styles["Title"]), [
            [Paragraph(row[0], styles['Heading3']), ReportGenerator.arabicParagraph(row[1], styles['Normal'])]
            for row in basicInfo
        ])

        def insertItemsTable(title, items: list[IC.IsolationItem]):
            """Append a titled table of isolation items (No./Tag/Description/State/
            Lock#/Lock Box#) to `elements`, followed by a page break. No-op if
            items is empty."""
            if not items:
                return

            nonlocal elements
            headers = ['No.', 'Tag', 'Description', 'State', 'Lock#', 'Lock Box#']
            weights = [7, 16, 40, 11, 11, 11]
            rows = [[Paragraph(h, styles['Heading3Center']) for h in headers]]
            for i, item in enumerate(items, start=1):
                rows.append([
                    Paragraph(str(i), styles['NormalCenter']),
                    ReportGenerator.arabicParagraph(item.tag or '', styles['Normal']),
                    ReportGenerator.arabicParagraph(item.description or '', styles['Normal']),
                    Paragraph(_html.escape(item.state or ''), styles['NormalCenter']),
                    ReportGenerator.arabicParagraph(item.lock_num or '', styles['NormalCenter'], forceAlignment=False),
                    ReportGenerator.arabicParagraph(item.lock_box_num or '', styles['NormalCenter'], forceAlignment=False),
                ])
            table = Table(rows, colWidths=[w * dataTableWidth / sum(weights) for w in weights], repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0, 0, 0, 0.2)),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.extend([title, Spacer(1, 0.2 * inch), table, PageBreak()])

        insertItemsTable(Paragraph('Isolation Items:', styles["Title"]), ic.items)

        if ic.linked_ptws:
            elements.extend([
                Paragraph('Linked PTWs:', styles["Title"]),
                Spacer(1, 0.2 * inch),
                listToBullets([f'PTW #{ptwId}' for ptwId in ic.linked_ptws], styles['Normal']),
                PageBreak(),
            ])

        try:
            pdfmetrics.registerFont(TTFont('Satisfy', resource_path('fonts/Satisfy/Satisfy-Regular.ttf')))
            sig_font = 'Satisfy'
        except Exception:
            sig_font = 'Helvetica-Oblique'

        label_style = ParagraphStyle('SigLabel', parent=styles['Heading3'], fontSize=13, leading=14, alignment=TA_CENTER)
        sig_style   = ParagraphStyle('Signature', parent=styles['Normal'],  fontSize=16, leading=17, alignment=TA_CENTER, fontName=sig_font)
        date_style  = ParagraphStyle('SigDate',   parent=styles['Normal'],  fontSize=12,  leading=14, alignment=TA_CENTER)

        def sigTables(columns: list, chunk_size=3):
            """columns: list of (label, sig, ts) or None for empty padding slots."""
            chunks = [columns[i:i+chunk_size] for i in range(0, len(columns), chunk_size)]
            result = []
            for chunk in chunks:
                col_w = dataTableWidth / len(chunk)
                header_row, name_row, date_row = [], [], []
                style_cmds = [
                    ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING',    (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]
                for i, col in enumerate(chunk):
                    if col is not None:
                        label, sig, ts = col
                        style_cmds.append(('LINEBELOW', (i, 0), (i, 0), 0.5, colors.black))
                        style_cmds.append(('LINEBELOW', (i, 1), (i, 1), 0.5, colors.black))
                    else:
                        label, sig, ts = '', '', ''
                    header_row.append(Paragraph(label or ' ',        label_style))
                    name_row.append(  ReportGenerator.arabicParagraph(sig or ' ', sig_style, forceAlignment=False))
                    date_row.append(  Paragraph(ts    or ' ',        date_style))
                table = Table(
                    [header_row, name_row, date_row],
                    colWidths=[col_w] * len(chunk),
                )
                table.setStyle(TableStyle(style_cmds))
                result.append(table)
                result.append(Spacer(1, 0.09 * inch))
            return result

        def approvalColumns():
            """Build one signature column per required approver stage/slot, padded
            to 5 with blanks: (approver label, approver's name if approved, else
            '', approval timestamp if approved, else '')."""
            def lastApprovalFor(approver):
                """Return the most recent approval on ic matching approver, or None."""
                match = None
                for approval in ic.approvals:
                    if approver.matchesUser(globalData.allUsers.get(approval.username)):
                        match = approval
                return match

            cols = []
            for stage in ic.requiredApprovers():
                for approver in stage:
                    approval = lastApprovalFor(approver)
                    if approval and approval.action == IC.ApprovalActions.APPROVED:
                        name = nameFor(approval.username)
                        ts   = approval.timestamp or ''
                    else:
                        name, ts = '', ''
                    cols.append((str(approver), name, ts))
            while len(cols) < 5:
                cols.append(None)
            return cols

        elements.extend([
            Paragraph('Approvals:', styles["Title"]),
            Spacer(1, 0.07 * inch),
            *sigTables(approvalColumns()),
            PageBreak(),
        ])

        isolateCols = [
            ('Isolate Requested', nameFor(ic.isolate_requestor), ic.isolate_requestor_timestamp or ''),
            (f'Isolate {ic.isolate_issuing_action}' if ic.isolate_issuing_action else 'Isolate Confirmed', nameFor(ic.isolate_issuing), ic.isolate_issuing_timestamp or ''),
            ('Isolate Carried Out', nameFor(ic.isolate_isolator), ic.isolate_isolator_timestamp or ''),
        ]
        deisolateCols = [
            ('De-isolate Requested', nameFor(ic.deisolate_requestor), ic.deisolate_requestor_timestamp or ''),
            (f'De-isolate {ic.deisolate_issuing_action}' if ic.deisolate_issuing_action else 'De-isolate Confirmed', nameFor(ic.deisolate_issuing), ic.deisolate_issuing_timestamp or ''),
            ('De-isolate Carried Out', nameFor(ic.deisolate_isolator), ic.deisolate_isolator_timestamp or ''),
        ]

        elements.extend([
            Paragraph('Isolation:', styles["Title"]),
            Spacer(1, 0.07 * inch),
            *sigTables(isolateCols),
        ])

        sanctionCols = [
            ('Sanction Requested', nameFor(ic.sanction_requestor), ic.sanction_requestor_timestamp or ''),
            ('Sanction Confirmed',  nameFor(ic.sanction_issuing),   ic.sanction_issuing_timestamp or ''),
            ('Sanction Carried Out', nameFor(ic.sanction_isolator), ic.sanction_isolator_timestamp or ''),
        ]
        elements.extend([
            Spacer(1, 0.1 * inch),
            Paragraph('Sanction for Test:', styles["Title"]),
            Spacer(1, 0.07 * inch),
            *sigTables(sanctionCols),
        ])

        reisolateCols = [
            ('Re-isolate Requested', nameFor(ic.reisolate_requestor), ic.reisolate_requestor_timestamp or ''),
            ('Re-isolate Confirmed',  nameFor(ic.reisolate_issuing),   ic.reisolate_issuing_timestamp or ''),
            ('Re-isolate Carried Out', nameFor(ic.reisolate_isolator), ic.reisolate_isolator_timestamp or ''),
        ]
        elements.extend([
            Spacer(1, 0.1 * inch),
            Paragraph('Re-isolation:', styles["Title"]),
            Spacer(1, 0.07 * inch),
            *sigTables(reisolateCols),
        ])

        elements.extend([
            Spacer(1, 0.1 * inch),
            Paragraph('De-isolation:', styles["Title"]),
            Spacer(1, 0.07 * inch),
            *sigTables(deisolateCols),
        ])

        class NumberedCanvas(canvas.Canvas):
            """ReportLab canvas that defers page-number drawing until save(), so
            each page can display "Page X of Y" with the final total page count."""

            def __init__(self, *args, **kwargs):
                """Initialize the underlying canvas and the buffer of saved page states."""
                canvas.Canvas.__init__(self, *args, **kwargs)
                self._saved_page_states = []

            def showPage(self):
                """Stash the current page's drawing state instead of finalizing the
                page immediately, so the page-number footer can be added later."""
                self._saved_page_states.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                """Replay each saved page state, draw its "Page X of Y" footer, then
                finalize all pages and save the document."""
                num_pages = len(self._saved_page_states)
                for state in self._saved_page_states:
                    self.__dict__.update(state)
                    self.setFont('Helvetica', 14)
                    self.setFillColorRGB(0, 0, 0, 1)
                    self.drawCentredString(0.7 * MARGIN + QR_CODE_WIDTH / 2, MARGIN, f'Page {self._pageNumber} of {num_pages}')
                    canvas.Canvas.showPage(self)
                canvas.Canvas.save(self)

        def pageHeaderAndWatermark(canvas: canvas.Canvas, doc):
            """Draw the two company logos, a diagonal "Printed @ <timestamp>"
            watermark, and the logo QR code into the page's left margin. Used as
            the ReportLab onFirstPage/onLaterPages callback."""
            canvas.saveState()

            canvas.drawImage(resource_path("assets/rashpetco-logo.png"), 0.7 * MARGIN + (QR_CODE_WIDTH - LOGO_IMG_WIDTH) / 2, (pageHeight - LOGO_IMG_WIDTH) / 2.0 + pageHeight / 3.5, LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, mask='auto')
            canvas.drawImage(resource_path("assets/burullus-logo.png"),  0.7 * MARGIN + (QR_CODE_WIDTH - LOGO_IMG_WIDTH) / 2, (pageHeight - LOGO_IMG_WIDTH) / 2.0 - pageHeight / 3.5, LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, mask='auto')

            canvas.setFont('Helvetica-Bold', 50)
            canvas.setFillColorRGB(0, 0, 0, 0.2)
            canvas.translate(pageWidth / 2.0, pageHeight / 2.0)
            canvas.rotate(35)
            canvas.drawCentredString(0, 0, f'Printed @ {timestamp}')

            canvas.restoreState()
            canvas.drawImage(qrPath, 0.7 * MARGIN, (pageHeight - QR_CODE_WIDTH) / 2.0, QR_CODE_WIDTH, QR_CODE_WIDTH)

        doc.build(elements, onFirstPage=pageHeaderAndWatermark, onLaterPages=pageHeaderAndWatermark, canvasmaker=NumberedCanvas)
        buffer.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, prefix=f'ic-{ic.id}-', suffix='.pdf') as icPdfFile:
            icPdfFile.write(buffer.read())
            icPdfFile.flush()
            ReportGenerator.openPDF(icPdfFile.name)

        return None


    def openPDF(filepath: str):
        """Open the given file in a new browser tab."""
        webbrowser.open_new_tab(filepath)
        return

        if platform.system() == "Windows":
            os.startfile(filepath)
        elif platform.system() == "Darwin":  # macOS
            subprocess.call(["open", filepath])
        else:  # Linux
            subprocess.call(["xdg-open", filepath])

        # webbrowser.open_new(filepath)
        # absPath = os.path.abspath(filepath)
        # url = 'file://' + absPath
        # try:
        #     from urllib.request import pathname2url
        # except ImportError:
        #     from urllib import pathname2url
        # url = 'file:{}'.format(pathname2url(absPath))
        # webbrowser.open_new_tab(url)


    def exportPTWs(ptws: list[PTW]):
        """Export a list of PTWs to an Excel workbook and open it.

        Writes one worksheet 'PTWs' with a styled header row (PTW#, Type, Status
        via `runningStatusDisplay()`, Date, Department, Requestor, PA, Location,
        Area Class, Equipment, Description) and one data row per PTW, each row
        filled and colored using `ptw.backgroundColor()`/`ptw.foregroundColor()`.
        Freezes the header row and enables auto-filter, then saves to a temp
        .xlsx file and opens it with the OS's default handler.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'PTWs'

        
        headers = ['PTW#', 'Type', 'Status', 'Date', 'Department', 'Requestor', 'PA', 'Location', 'Area Class', 'Equipment', 'Description']
        col_widths = [8, 20, 15, 14, 18, 22, 22, 20, 12, 25, 50]

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='2E4057')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='AAAAAA')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 28

        def qcolor_to_hex(qcolor):
            """Convert a QColor to an RRGGBB hex string, blending its RGB channels
            toward white using a fixed alpha of 240 (the QColor's own alpha is
            ignored)."""
            r, g, b, a = qcolor.red(), qcolor.green(), qcolor.blue(), 240
            if a < 255:
                r = (r * a + 255 * (255 - a)) // 255
                g = (g * a + 255 * (255 - a)) // 255
                b = (b * a + 255 * (255 - a)) // 255
            return f'{r:02X}{g:02X}{b:02X}'

        cell_align = Alignment(vertical='center', wrap_text=True)
        # Arabic cell values (department/requestor/location/equipment/description, etc.)
        # read right-to-left, so right-align just those cells rather than the whole
        # column - openpyxl/Excel already renders Arabic glyphs natively, no font or
        # bidi-reordering fix is needed here the way it is for the PDF reports.
        cell_align_rtl = Alignment(horizontal='right', vertical='center', wrap_text=True)

        for row_idx, ptw in enumerate(ptws, start=2):
            status = ptw.runningStatusDisplay()
            requestor = globalData.allUsers[ptw.requestor].getName() if ptw.requestor in globalData.allUsers else str(ptw.requestor or '')
            performing = ptw.getPerforming()
            pa = globalData.allUsers[performing].getName() if performing in globalData.allUsers else str(performing or '')

            row_data = [
                ptw.id,
                str(ptw.type or ''),
                status,
                str(ptw.request_date or ''),
                str(ptw.department or ''),
                requestor,
                pa,
                str(ptw.location or ''),
                str(ptw.area_class or ''),
                str(ptw.equipment or ''),
                str(ptw.description or ''),
            ]

            row_fill = PatternFill(fill_type='solid', fgColor=qcolor_to_hex(ptw.backgroundColor()))
            row_font = Font(color=qcolor_to_hex(ptw.foregroundColor()))
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_align_rtl if isRtlBase(str(value)) else cell_align
                cell.border = border
                cell.fill = row_fill
                cell.font = row_font

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        timestamp = datetime.now().strftime('%d-%m-%Y_%H-%M-%S')
        with tempfile.NamedTemporaryFile(delete=False, prefix=f'ptws-export-{timestamp}-', suffix='.xlsx') as f:
            wb.save(f.name)
            if platform.system() == 'Windows':
                os.startfile(f.name)
            elif platform.system() == 'Darwin':
                subprocess.call(['open', f.name])
            else:
                subprocess.call(['xdg-open', f.name])


    def MOSReport(mos: str, ptwId: str, ptwTitle: str):
        """Build and open a standalone Method of Statement PDF for a PTW.

        Lays out a portrait-A4 document whose header (on every page) shows the two
        company logos flanking a "MOS for PTW# <id>" title and a diagonal "Printed
        @ <timestamp>" watermark; the body is `mos` split on newlines and rendered
        as a bulleted list. No-op if `mos` is empty.
        """
        ReportGenerator._registerArabicFonts()

        LOGO_IMG_WIDTH = 30*mm
        MARGIN = 0.8 * inch

        if not mos:
            return

        buffer = io.BytesIO()
        timestamp = datetime.now().strftime('%d-%m-%Y - %H:%M:%S')
        doc = SimpleDocTemplate(buffer, pagesize=portrait(A4), leftMargin=MARGIN, rightMargin=MARGIN, topMargin=1.5*MARGIN+LOGO_IMG_WIDTH, bottomMargin=MARGIN)

        pageWidth, pageHeight = portrait(A4)
        headerTableWidth = pageWidth - 2*MARGIN

        styles = getSampleStyleSheet()
        
        styles['Title'].fontSize = 20
        styles['Title'].leading = 20
        styles["Title"].alignment = TA_CENTER

        styles['Normal'].fontSize = 16
        styles['Normal'].leading = 16
        styles['Normal'].fontName = 'Helvetica'

        styles['Heading3'].fontSize = 18
        styles['Heading3'].leading = 18
        styles['Heading3'].fontName = 'Helvetica-Bold'
        styles['Heading3'].alignment = TA_CENTER

        def pageHeaderAndWatermark(canvas: canvas.Canvas, doc):
            """Draw the header table (two company logos flanking a "MOS for
            PTW# <id>" title) and a diagonal "Printed @ <timestamp>" watermark.
            Used as the ReportLab onFirstPage/onLaterPages callback."""
            canvas.saveState()

            logo1 = Image(resource_path("assets/rashpetco-logo.png"), LOGO_IMG_WIDTH, LOGO_IMG_WIDTH)
            logo2 = Image(resource_path("assets/burullus-logo.png"),  LOGO_IMG_WIDTH, LOGO_IMG_WIDTH)
            label = Paragraph(f'MOS for PTW# {ptwId} <br/>' + pdfMarkup(ptwTitle), styles['Title'])

            table = Table([[logo1, label, logo2]], colWidths=[1.1*LOGO_IMG_WIDTH, headerTableWidth - 2.2*LOGO_IMG_WIDTH, 1.1*LOGO_IMG_WIDTH])
            table.setStyle(TableStyle([
                # Alignment
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), 
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), 

                # grid (optional)
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            table.wrapOn(canvas, headerTableWidth, LOGO_IMG_WIDTH*1.2)
            table.drawOn(canvas, MARGIN, pageHeight - MARGIN - LOGO_IMG_WIDTH)

            canvas.setFont('Helvetica-Bold', 50)
            canvas.setFillColorRGB(0, 0, 0, 0.2)
            canvas.translate(pageWidth / 2.0, pageHeight / 2.0)
            canvas.rotate(55)
            canvas.drawCentredString(0, 0, f'Printed @ {timestamp}')

            canvas.restoreState()
        
        def listToBullets(data: list, style):
            """Render a list of strings as a bulleted ListFlowable of Arabic-aware
            paragraphs (see ReportGenerator.arabicParagraph), or None if data is empty."""
            if not data:
                return None

            bullets = [ListItem(ReportGenerator.arabicParagraph(element, style), bulletType='bullet', bulletFontSize=style.fontSize) for element in data]
            return ListFlowable(
                bullets,
                bulletType='bullet',
                leftIndent=0.3*inch,
                bulletIndent=0.1*inch,
                spaceAfter=0.1*inch,
            )

        elements = []
        mosSteps = mos.split('\n')
        elements.append(Paragraph('Method of Statement:', styles['Title']))
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(listToBullets(mosSteps, styles['Normal']))

        doc.build(elements, onFirstPage=pageHeaderAndWatermark, onLaterPages=pageHeaderAndWatermark)
        buffer.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, prefix=f'mos-{ptwId}-', suffix='.pdf') as mosPdfFile:
            mosPdfFile.write(buffer.read())
            mosPdfFile.flush()
            ReportGenerator.openPDF(mosPdfFile.name)
        

    def riskAssessmentReport(riskAssessment: RiskAssessment = None):
        """Build and open a risk assessment PDF, laid out as a landscape-A4 table.

        The table has one row per risk item with columns No./Hazard/Effect, a
        grouped "Free Analysis" (Severity/Likelihood/Risk) and "Controlled
        Analysis" (Severity/Likelihood/Risk) pair computed from each item's
        `free_analysis`/`ctrl_analysis` strings, Control, and Evaluation; hazard,
        effect, and control text are rendered as bulleted lines split on newlines.
        The page header (on every page) shows the two company logos flanking the
        assessment's title and a diagonal "Printed @ <timestamp>" watermark.
        No-op if riskAssessment is None or has no risks.
        """
        ReportGenerator._registerArabicFonts()

        LOGO_IMG_WIDTH = 35*mm
        # Cols: No | Hazard | Effect | S | L | Risk(free) | Control | S | L | Risk(ctrl) | Evaluation
        TABLE_WIDTH_WEIGHTS = [7, 30, 33, 4, 4, 9, 52, 5, 5, 10, 20]
        TABLE_WIDTH_WEIGHTS_SUM = sum(TABLE_WIDTH_WEIGHTS)
        MARGIN = 0.35 * inch

        if not riskAssessment or not riskAssessment.risks:
            return

        ptwId = riskAssessment.ptw_id
        title = riskAssessment.title
        
        buffer = io.BytesIO()
        timestamp = datetime.now().strftime('%d-%m-%Y - %H:%M:%S')
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=2*MARGIN, rightMargin=2*MARGIN, topMargin=1.5*MARGIN+LOGO_IMG_WIDTH, bottomMargin=MARGIN)

        pageWidth, pageHeight = landscape(A4)
        dataTableWidth = pageWidth - 4*MARGIN

        styles = getSampleStyleSheet()
        
        styles['Title'].fontSize = 20
        styles['Title'].leading = 20
        styles["Title"].alignment = TA_CENTER

        styles['Normal'].fontSize = 12
        styles['Normal'].leading = 12
        styles['Normal'].fontName = 'Helvetica'

        styles['Heading3'].fontSize = 14
        styles['Heading3'].leading = 14
        styles['Heading3'].fontName = 'Helvetica-Bold'
        styles['Heading3'].alignment = TA_CENTER

        styles.add(ParagraphStyle('NormalCenter',  parent=styles['Normal'], alignment=TA_CENTER))
        styles.add(ParagraphStyle('NormalJustify', parent=styles['Normal'], alignment=TA_LEFT))
        _bw = pdfmetrics.stringWidth('• ', 'Helvetica', 12)
        styles.add(ParagraphStyle('BulletItem',    parent=styles['NormalJustify'], leftIndent=_bw, firstLineIndent=-_bw))

        def slr(text):
            """Split a Severity/Likelihood/Risk analysis string into its (severity,
            likelihood, full text) parts: the first and last characters of text,
            plus text itself."""
            t = text or ''
            return t[0], t[-1], t

        def bulleted(text):
            """Split text on newlines, dropping blank lines, and render each
            remaining line as a bulleted, Arabic-aware paragraph (see
            ReportGenerator.arabicParagraph)."""
            lines = [l for l in (text or '').split('\n') if l.strip()]
            return [ReportGenerator.arabicParagraph('• ' + line, styles['BulletItem']) for line in lines] if lines else []

        data = [
            # Row 0 — group headers (non-split cols span both rows via SPAN)
            [
                Paragraph('No.',                  styles['Heading3']), 
                Paragraph('Hazard',               styles['Heading3']), 
                Paragraph('Effect',               styles['Heading3']),
                Paragraph('Free Analysis',        styles['Heading3']), 
                Paragraph('',                     styles['Heading3']), 
                Paragraph('',                     styles['Heading3']),
                Paragraph('Control',              styles['Heading3']),
                Paragraph('Controlled Analysis',  styles['Heading3']), 
                Paragraph('',                     styles['Heading3']), 
                Paragraph('',                     styles['Heading3']),
                Paragraph('Evaluation',           styles['Heading3']), 
            ],
            # Row 1 — sub-headers for split cols
            [
                Paragraph('',                     styles['Heading3']), 
                Paragraph('',                     styles['Heading3']), 
                Paragraph('',                     styles['Heading3']),
                Paragraph('S',                    styles['Heading3']), 
                Paragraph('L',                    styles['Heading3']), 
                Paragraph('Risk',                 styles['Heading3']),
                Paragraph('',                     styles['Heading3']),
                Paragraph('S',                    styles['Heading3']), 
                Paragraph('L',                    styles['Heading3']), 
                Paragraph('Risk',                 styles['Heading3']),
                Paragraph('',                     styles['Heading3']), 
            ],
        ]
        for riskItem in riskAssessment.risks:
            sf, lf, rf = slr(riskItem.free_analysis)
            sc, lc, rc = slr(riskItem.ctrl_analysis)
            data.append([
                Paragraph(str(len(data) - 1),           styles['NormalCenter']),
                bulleted(riskItem.hazard),
                bulleted(riskItem.effect),
                Paragraph(sf,                           styles['NormalCenter']),
                Paragraph(lf,                           styles['NormalCenter']),
                Paragraph(rf,                           styles['NormalCenter']),
                bulleted(riskItem.ctrl),
                Paragraph(sc,                           styles['NormalCenter']),
                Paragraph(lc,                           styles['NormalCenter']),
                Paragraph(rc,                           styles['NormalCenter']),
                ReportGenerator.arabicParagraph(riskItem.eval, styles['NormalCenter'], forceAlignment=False),
            ])

        table = Table(data, repeatRows=2, colWidths=[w * dataTableWidth / TABLE_WIDTH_WEIGHTS_SUM for w in TABLE_WIDTH_WEIGHTS])
        table.setStyle(TableStyle([
            # header background (both rows)
            ('BACKGROUND', (0, 0), (-1, 1), colors.Color(0, 0, 0, 0.2)),

            # rowspan for non-split header cols
            ('SPAN', (0,  0), (0,  1)),   # No.
            ('SPAN', (1,  0), (1,  1)),   # Hazard
            ('SPAN', (2,  0), (2,  1)),   # Effect
            ('SPAN', (6,  0), (6,  1)),   # Control
            ('SPAN', (10, 0), (10, 1)),   # Evaluation

            # colspan for group headers
            ('SPAN', (3, 0), (5, 0)),     # Free Analysis
            ('SPAN', (7, 0), (9, 0)),     # Controlled Analysis

            # Alignment
            ('ALIGN',  (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

            # padding
            ('TOPPADDING',    (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING',   (0, 0), (-1, -1), 3),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 3),

            # grid
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements = [table]
            
        def pageHeaderAndWatermark(canvas: canvas.Canvas, doc):
            """Draw the header table (two company logos flanking the risk
            assessment's title) and a diagonal "Printed @ <timestamp>" watermark.
            Used as the ReportLab onFirstPage/onLaterPages callback."""
            canvas.saveState()

            # canvas.drawImage("./rashpetco-logo.png", MARGIN + 2.0 * pageWidth / 3.0, pageHeight - MARGIN - LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, mask='auto')
            # canvas.drawImage("./burullus-logo.png",  MARGIN + 1.0 * pageWidth / 3.0, pageHeight - MARGIN - LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, mask='auto')

            logo1 = Image(resource_path("assets/rashpetco-logo.png"), LOGO_IMG_WIDTH, LOGO_IMG_WIDTH)
            logo2 = Image(resource_path("assets/burullus-logo.png"),  LOGO_IMG_WIDTH, LOGO_IMG_WIDTH)
            label = Paragraph(
                (f'PTW#{ptwId} - Specific Risk Assessment <br/>' if ptwId else '') + pdfMarkup(title),
                styles['Title']
            )
            print(ptwId)
            print(title)

            table = Table([[logo1, label, logo2]], colWidths=[1.2*LOGO_IMG_WIDTH, dataTableWidth - 2.4*LOGO_IMG_WIDTH, 1.2*LOGO_IMG_WIDTH])
            table.setStyle(TableStyle([
                # Alignment
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), 
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), 

                # grid (optional)
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            table.wrapOn(canvas, dataTableWidth, LOGO_IMG_WIDTH*1.2)
            table.drawOn(canvas, 2*MARGIN, pageHeight - MARGIN - LOGO_IMG_WIDTH)

            canvas.setFont('Helvetica-Bold', 50)
            canvas.setFillColorRGB(0, 0, 0, 0.2)
            canvas.translate(pageWidth / 2.0, pageHeight / 2.0)
            canvas.rotate(35)
            canvas.drawCentredString(0, 0, f'Printed @ {timestamp}')

            canvas.restoreState()
        
        doc.build(elements, onFirstPage=pageHeaderAndWatermark, onLaterPages=pageHeaderAndWatermark)
        buffer.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, prefix=f'risk-{ptwId}-', suffix='.pdf') as riskPdfFile:
            riskPdfFile.write(buffer.read())
            riskPdfFile.flush()
            ReportGenerator.openPDF(riskPdfFile.name)
        
