"""Bulk user import from Excel/CSV files.

Parses a spreadsheet of user records into validated rows (`ImportUsersExcel.parseFile`),
presents them for review/confirmation and later results in a read-only table dialog
(`DialogUsersPreview`), and can export the outcome of an import back to Excel
(`ImportUsersExcel.exportResult`).
"""

import secrets
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem,
                              QAbstractItemView, QHeaderView, QDialogButtonBox)
from models.User import User, UserRoles, UserDepartments
from helper.utils import parseTabularFile


class DialogUsersPreview(QDialog):
    """Read-only table dialog previewing import rows before import, or showing
    per-row results afterward with an optional Excel export button."""

    def __init__(self, parent, title: str, headers: list[str], rows: list[list],
                 mode: str = 'confirm', summary: str = '', onExport=None):
        """Build the read-only rows table and its button row.

        Args:
            headers: column header labels.
            rows: table rows, each a list of stringified cell values.
            mode: 'confirm' shows Import/Cancel buttons for approving the import;
                any other value shows a Close button plus, if `onExport` is given,
                an "Export as Excel" button.
            summary: optional summary label shown above the table.
            onExport: callback invoked when "Export as Excel" is clicked.
        """
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(750, 400)

        lyt = QVBoxLayout()
        self.setLayout(lyt)

        if summary:
            lblSummary = QLabel(summary)
            lblSummary.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lyt.addWidget(lblSummary)

        self.tbl = QTableWidget()
        self.tbl.setColumnCount(len(headers))
        self.tbl.setHorizontalHeaderLabels(headers)
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.tbl.horizontalHeader().setStretchLastSection(True)
        self.tbl.verticalHeader().hide()

        self.tbl.setRowCount(len(rows))
        for r, record in enumerate(rows):
            for c, value in enumerate(record):
                self.tbl.setItem(r, c, QTableWidgetItem(str(value)))
        lyt.addWidget(self.tbl)

        btns = QDialogButtonBox()
        if mode == 'confirm':
            btnOk = btns.addButton("Import", QDialogButtonBox.ButtonRole.AcceptRole)
            btnCancel = btns.addButton(QDialogButtonBox.StandardButton.Cancel)
            btnOk.clicked.connect(self.accept)
            btnCancel.clicked.connect(self.reject)
        else:
            if onExport:
                btnExport = btns.addButton("Export as Excel", QDialogButtonBox.ButtonRole.ActionRole)
                btnExport.clicked.connect(onExport)
            btnClose = btns.addButton(QDialogButtonBox.StandardButton.Close)
            btnClose.clicked.connect(self.accept)
        lyt.addWidget(btns)


HEADERS = ['Username', 'Name', 'Role', 'Department', 'Email', 'EXT']
FIELDS = ['username', 'name', 'role', 'department', 'email', 'ext']


class ImportRow:
    """One parsed row from a bulk-import file, its validation status, and the
    resulting User object when the row was valid."""

    def __init__(self, rowNum: int, username: str, name: str, role: str, department: str,
                 email: str, ext: str, user: User = None, status: str = ''):
        """Store the parsed field values plus the constructed User (if valid) and status."""
        self.rowNum = rowNum
        self.username = username
        self.name = name
        self.role = role
        self.department = department
        self.email = email
        self.ext = ext
        self.user = user
        self.status = status

    def asRecord(self) -> list:
        """Return this row's values as a flat list for display/export.

        Returns:
            list: username, name, role, department, email, ext, the generated
            password (empty string if the row failed validation), and status.
        """
        return [
            self.username, self.name, self.role, self.department, self.email, self.ext,
            self.user.getPassword() if self.user else '',
            self.status,
        ]


class ImportUsersExcel:
    """Parses bulk user import spreadsheets and exports import results back to Excel."""

    HEADERS = HEADERS
    FIELDS = FIELDS

    @staticmethod
    def parseFile(filepath: str, existingUsernames: set) -> list[ImportRow]:
        """Parse a users spreadsheet and validate each row.

        Reads rows via `parseTabularFile` (matching HEADERS regardless of the file's
        column order or header casing/whitespace) and skips rows that are entirely
        blank. Each remaining row is rejected (with a descriptive status) if the
        username is empty or already taken (by `existingUsernames` or a duplicate
        earlier in the file), the name is empty, or the role/department don't match
        a `UserRoles`/`UserDepartments` value (case-insensitively). Valid rows get a
        new `User` built with a freshly generated random password.

        Args:
            filepath: path to the .xlsx or .csv file to import.
            existingUsernames: usernames already present in the system.

        Returns:
            list[ImportRow]: one entry per non-blank row, in file order.
        """
        dataRows = parseTabularFile(filepath, HEADERS)

        roleValues = {r.value.lower(): r.value for r in UserRoles}
        deptValues = {d.value.lower(): d.value for d in UserDepartments}

        rows = []
        seenUsernames = set()

        for rowNum, record in enumerate(dataRows, start=2):
            if all(not v for v in record):
                continue

            data = dict(zip(FIELDS, record))

            username = data['username']
            name = data['name']
            role = data['role']
            department = data['department']

            error = None
            if not username:
                error = "Username can't be empty"
            elif username in existingUsernames or username in seenUsernames:
                error = f"Username '{username}' already exists"
            elif not name:
                error = "Name can't be empty"
            elif role.lower() not in roleValues:
                error = f"Invalid role '{role}'"
            elif department.lower() not in deptValues:
                error = f"Invalid department '{department}'"

            user = None
            if error is None:
                seenUsernames.add(username)
                user = User(
                    username=username,
                    password=secrets.token_urlsafe(12),
                    name=name,
                    role=roleValues[role.lower()],
                    department=deptValues[department.lower()],
                    email=data['email'],
                )
                user.setExt(data['ext'])

            rows.append(ImportRow(
                rowNum=rowNum,
                username=username,
                name=name,
                role=role,
                department=department,
                email=data['email'],
                ext=data['ext'],
                user=user,
                status="Ready to import" if user else f"Skipped: {error}",
            ))

        return rows

    @staticmethod
    def exportResult(filepath: str, rows: list[ImportRow]):
        """Write the given import rows to an Excel workbook.

        Args:
            filepath: destination .xlsx path.
            rows: rows to write, one spreadsheet row each, under a header row of
                HEADERS + ['Password', 'Status'].
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Import Result'

        headers = HEADERS + ['Password', 'Status']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        for rowIdx, row in enumerate(rows, start=2):
            for col, value in enumerate(row.asRecord(), start=1):
                ws.cell(row=rowIdx, column=col, value=value)

        wb.save(filepath)
