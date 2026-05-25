import html as _html
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import *
from reportlab.lib import colors
from reportlab.lib.styles import *
from reportlab.lib.enums import *
from reportlab.platypus import *
from reportlab.lib.units import *
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import qrcode
from PIL import Image as PILImage, ImageDraw as PILImageDraw
import webbrowser
from datetime import datetime
from typing import Iterable
from GlobalData import globalData
from PTWData import PTWData, Isolation
import io
import tempfile
import platform
import os
import subprocess
from clientRequests import ClientRequests
from pypdf import PdfWriter, PdfReader


class ReportGenerator:

    def _makeQrWithLogo(ptw: PTWData):
        basicInfo = [
            ['PTW#', str(ptw.id)],
            ['Type', str(ptw.type)],
            ['Status', str(ptw.running_status if ptw.approval_status == PTWData.ApprovalStatus.APPROVED and ptw.running_status is not None else ptw.approval_status)],
            ['Department', str(ptw.department)],
            ['Requestor', str(globalData.allUsers[ptw.requestor].getName()) if ptw.requestor in globalData.allUsers else 'None'],
            ['PA', str(globalData.allUsers[ptw.performing].getName()) if ptw.performing in globalData.allUsers else 'None'],
            ['Location', str(ptw.location)],
            ['Equipment', str(ptw.equipment)],
            ['Description', str(ptw.description)],
        ]

        data = '\n'.join([': '.join(row) for row in basicInfo])
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_Q, box_size=20, border=2)
        qr.add_data(data)
        qr.make(fit=True)
        qrImg = qr.make_image(fill_color="black", back_color="white").convert("RGB")
        qr_w, qr_h = qrImg.size

        logo_size = qr_w // 3
        padding = logo_size // 32
        logo = PILImage.open("./sh-logo-bw.png").convert("RGB").resize((logo_size, logo_size), PILImage.LANCZOS)
        logo_mask = PILImage.new("L", (logo_size, logo_size), 0)
        PILImageDraw.Draw(logo_mask).ellipse((0, 0, logo_size - 1, logo_size - 1), fill=255)
        total_size = logo_size + 2 * padding
        bg = PILImage.new("RGB", (total_size, total_size), "white")
        bg_mask = PILImage.new("L", (total_size, total_size), 0)
        PILImageDraw.Draw(bg_mask).ellipse((0, 0, total_size - 1, total_size - 1), fill=255)
        bg.paste(logo, (padding, padding), logo_mask)
        qrImg.paste(bg, ((qr_w - total_size) // 2, (qr_h - total_size) // 2), bg_mask)
        qrImg.save(f"qr-summery-{ptw.id}.png")
        

    def ptwReport(loggedUser, ptw: PTWData):
        QR_CODE_WIDTH = 60*mm
        LOGO_IMG_WIDTH = 40*mm
        TABLE_LABELS_WIDTH = 50*mm
        MARGIN = 0.6 * inch

        buffer = io.BytesIO()
        timestamp = datetime.now().strftime('%d-%m-%Y - %H:%M:%S')
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=MARGIN+QR_CODE_WIDTH, rightMargin=MARGIN, topMargin=MARGIN*0.7, bottomMargin=MARGIN*0.7)

        pageWidth, pageHeight = landscape(A4)
        dataTableWidth = pageWidth - 2*MARGIN - QR_CODE_WIDTH

        styles = getSampleStyleSheet()
        
        styles['Title'].fontSize = 24
        styles['Title'].leading = 26
        styles["Title"].alignment = TA_LEFT

        styles['Normal'].fontSize = 16
        styles['Normal'].leading = 18
        styles['Normal'].fontName = 'Helvetica'

        styles['Heading3'].fontSize = 16
        styles['Heading3'].leading = 18
        styles['Heading3'].fontName = 'Helvetica-Bold'

        def listTo2ColsBullets(data: list, style):
            mid = (len(data) + 1) // 2
            col1Data = data[:mid]
            col2Data = data[mid:]
            tableData = [
                [Paragraph('• ' + _html.escape(col1Data[i]), style), (Paragraph('• ' + _html.escape(col2Data[i]), style) if i < len(col2Data) else None)]
                for i in range(len(col1Data))
            ]
            return Table(tableData)

        def listToBullets(data: list, style):
            if not data:
                return None
            
            bullets = [ListItem(Paragraph(_html.escape(element), style), bulletType='bullet', bulletFontSize=style.fontSize) for element in data]
            return ListFlowable(
                bullets, 
                bulletType='bullet', 
                leftIndent=0.3*inch, 
                bulletIndent=0.1*inch, 
                spaceAfter=0.1*inch, 
            )

        basicInfo = [
            ['PTW#', str(ptw.id)], 
            ['Type', str(ptw.type)], 
            ['Print Time', timestamp], 
            ['Status', str(ptw.running_status if ptw.approval_status == PTWData.ApprovalStatus.APPROVED and ptw.running_status is not None else ptw.approval_status)], 
            ['Request Date', str(ptw.date)], 
            ['Department', str(ptw.department)], 
            ['Requestor', str(globalData.allUsers[ptw.requestor].getName()) if ptw.requestor in globalData.allUsers else 'None'], 
            ['PA', str(globalData.allUsers[ptw.performing].getName()) if ptw.performing in globalData.allUsers else 'None'], 
            ['Location', str(ptw.location)], 
            ['Equipment', str(ptw.equipment)], 
            ['Area Class', str(ptw.area_class)], 
            ['Description', str(ptw.description)], 
        ]

        ReportGenerator._makeQrWithLogo(ptw)

        elements = []

        def insertTable(title, tableData: list[list]):
            if not tableData:
                return
            
            nonlocal elements
            table = Table(tableData, colWidths=[TABLE_LABELS_WIDTH, dataTableWidth - TABLE_LABELS_WIDTH], splitByRow=1)
            table.setStyle(TableStyle([
                # header column (left)
                ('BACKGROUND', (0, 0), (0, -1), colors.Color(0, 0, 0, 0.2)),

                # Alignment
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'), 
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), 

                # padding
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),

                # grid (optional)
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.extend([title, Spacer(1, 0.2 * inch), table, PageBreak()])

        insertTable(Paragraph('Summery:', styles["Title"]), [
            [Paragraph(row[0], styles['Heading3']), Paragraph(_html.escape(row[1]), styles['Normal'])]
            for row in basicInfo
        ])

        tableAdditionalInfoData = [
            [Paragraph('Tools', styles['Heading3']), listToBullets(ptw.tools, styles['Normal'])], 
            [Paragraph('Hazards', styles['Heading3']), listToBullets(ptw.hazards, styles['Normal'])], 
            [Paragraph('Controls', styles['Heading3']), listToBullets(ptw.controls, styles['Normal'])], 
            [Paragraph('Risks', styles['Heading3']), listToBullets(ptw.risks, styles['Normal'])], 
        ]
        if ptw.miwi:
            tableAdditionalInfoData.append([Paragraph('MIWI', styles['Heading3']), listToBullets([ptw.miwi], styles['Normal'])])
        # elif ptw.mos:
        #     tableAdditionalInfoData.append([Paragraph('MOS', styles['Heading3']), listToBullets(ptw.mos.split('\n'), styles['Normal'])])
        if ptw.attachs:
            tableAdditionalInfoData.append([Paragraph('Attachments', styles['Heading3']), listToBullets(ptw.attachs, styles['Normal'])])
        # if ptw.isolations:
        #     tableAdditionalInfoData.append([Paragraph('Isolations', styles['Heading3']), listToBullets([str(isolation) for isolation in ptw.isolations], styles['Normal'])])

        insertTable(Paragraph('Additional Info:', styles["Title"]), tableAdditionalInfoData)


        if ptw.mos:
            mosSteps = ptw.mos.split('\n')
            elements.extend([Paragraph('Method Of Statement:', styles["Title"]), Spacer(1, 0.2 * inch)])
            elements.append(listToBullets(mosSteps, styles['Normal']))
            elements.append(PageBreak())
        

        isolationsByType = {}
        for iso in ptw.isolations:
            if iso.type not in isolationsByType:
                isolationsByType[iso.type] = []
            isolationsByType[iso.type].append(iso)
        
        if ptw.running_status == PTWData.RunningStatus.RUNNING:
            for isoType in isolationsByType:
                elements.extend([
                    Paragraph(f'{isoType} Isolations:', styles["Title"]),
                    Spacer(1, 0.07 * inch),
                    ReportGenerator._isolationsTable(ptw, isolationsByType[isoType], dataTableWidth, isolate=True), 
                    PageBreak(),
                ])
        # elif ptw.running_status == PTWData.RunningStatus.CLOSED:
        #     for isoType in isolationsByType:
        #         elements.extend([
        #             Paragraph(f'{isoType} De-Isolations:', styles["Title"]),
        #             Spacer(1, 0.07 * inch),
        #             ReportGenerator._isolationsTable(ptw, isolationsByType[isoType], dataTableWidth, isolate=False), 
        #             PageBreak(),
        #         ])

        try:
            pdfmetrics.registerFont(TTFont('Satisfy', './fonts/Satisfy/Satisfy-Regular.ttf'))
            sig_font = 'Satisfy'
        except Exception:
            sig_font = 'Helvetica-Oblique'

        label_style = ParagraphStyle('SigLabel', parent=styles['Heading3'], fontSize=16, leading=18, alignment=TA_CENTER)
        sig_style   = ParagraphStyle('Signature', parent=styles['Normal'],  fontSize=16, leading=18, alignment=TA_CENTER, fontName=sig_font)
        date_style  = ParagraphStyle('SigDate',   parent=styles['Normal'],  fontSize=14, leading=16, alignment=TA_CENTER)

        def sigTables(columns: list, chunk_size=3):
            """columns: list of (label, sig, ts) or None for empty padding slots."""
            chunks = [columns[i:i+chunk_size] for i in range(0, len(columns), chunk_size)]
            result = []
            for chunk in chunks:
                col_w = dataTableWidth / len(chunk)
                header_row, name_row, date_row = [], [], []
                style_cmds = [
                    ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING',    (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]
                for i, col in enumerate(chunk):
                    if col is not None:
                        label, sig, ts = col
                        style_cmds.append(('LINEBELOW', (i, 0), (i, 0), 0.5, colors.black))
                        style_cmds.append(('LINEBELOW', (i, 1), (i, 1), 0.5, colors.black))
                    else:
                        label, sig, ts = '', '', ''
                    header_row.append(Paragraph(label or '\u00a0',        label_style))
                    name_row.append(  Paragraph(sig   or '\u00a0',    sig_style))
                    date_row.append(  Paragraph(ts    or '\u00a0',        date_style))
                table = Table(
                    [header_row, name_row, date_row],
                    colWidths=[col_w] * len(chunk),
                )
                table.setStyle(TableStyle(style_cmds))
                result.append(table)
                result.append(Spacer(1, 0.2 * inch))
            return result

        def approvalColumns():
            role_approval: dict = {}
            for approval in ptw.approvals:
                if approval.username in globalData.allUsers:
                    role = globalData.allUsers[approval.username].getRole()
                    role_approval[role] = approval
            cols = []
            for role in ptw.requiredApprovers():
                approval = role_approval.get(role)
                if approval and approval.action == PTWData.ApprovalActions.APPROVED:
                    name = globalData.allUsers[approval.username].getName()
                    ts   = approval.timestamp or ''
                else:
                    name, ts = '', ''
                cols.append((str(role), name, ts))
            while len(cols) < 7:
                cols.append(None)
            return cols

        def runConfirmationColumns():
            pa = globalData.allUsers[ptw.performing].getName() if ptw.performing in globalData.allUsers else str(ptw.performing or '')
            ia = globalData.allUsers[ptw.issuing].getName()    if ptw.issuing    in globalData.allUsers else str(ptw.issuing    or '')
            return [
                ('PA', pa, ptw.performing_timestamp if ptw.performing_timestamp else ''),
                ('IA', ia, ptw.issuing_timestamp    if ptw.issuing_timestamp    else ''),
            ]

        elements.extend([
            Paragraph('Approvals:', styles["Title"]),
            Spacer(1, 0.07 * inch),
            *sigTables(approvalColumns()),
            Spacer(1, 0.1 * inch),
            Paragraph('Running Confirmations:', styles["Title"]),
            Spacer(1, 0.07 * inch),
            *sigTables(runConfirmationColumns()),
            PageBreak(),
        ])

        # insertTable(Paragraph('Approval Cycle:', styles["Title"]), [
        #     [Paragraph(approval.action, styles['Heading3']), Paragraph(str(approval), styles['Normal'])]
        #     for approval in ptw.approvals
        # ])
        

        class NumberedCanvas(canvas.Canvas):
            def __init__(self, *args, **kwargs):
                canvas.Canvas.__init__(self, *args, **kwargs)
                self._saved_page_states = []

            def showPage(self):
                self._saved_page_states.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                num_pages = len(self._saved_page_states)
                for state in self._saved_page_states:
                    self.__dict__.update(state)
                    self.setFont('Helvetica', 14)
                    self.setFillColorRGB(0, 0, 0, 1)
                    self.drawCentredString(0.7 * MARGIN + QR_CODE_WIDTH / 2, MARGIN, f'Page {self._pageNumber} of {num_pages}')
                    canvas.Canvas.showPage(self)
                canvas.Canvas.save(self)

        def pageHeaderAndWatermark(canvas: canvas.Canvas, doc):
            canvas.saveState()

            canvas.drawImage("./rashpetco-logo.png", 0.7 * MARGIN + (QR_CODE_WIDTH - LOGO_IMG_WIDTH) / 2, (pageHeight - LOGO_IMG_WIDTH) / 2.0 + pageHeight / 3.5, LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, mask='auto')
            canvas.drawImage("./burullus-logo.png",  0.7 * MARGIN + (QR_CODE_WIDTH - LOGO_IMG_WIDTH) / 2, (pageHeight - LOGO_IMG_WIDTH) / 2.0 - pageHeight / 3.5, LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, mask='auto')

            canvas.setFont('Helvetica-Bold', 50)
            canvas.setFillColorRGB(0, 0, 0, 0.2)
            canvas.translate(pageWidth / 2.0, pageHeight / 2.0)
            canvas.rotate(35)
            canvas.drawCentredString(0, 0, f'Printed @ {timestamp}')

            canvas.restoreState()
            canvas.drawImage(f"qr-summery-{ptw.id}.png", 0.7 * MARGIN, (pageHeight - QR_CODE_WIDTH) / 2.0, QR_CODE_WIDTH, QR_CODE_WIDTH)

        # toolsTitle = Paragraph('Tools:', styles["Title"])
        # toolsBullets = [ListItem(Paragraph(tool, styles['Normal']), bulletType='bullet', bulletFontSize=styles['Normal'].fontSize) for tool in ptw.tools]
        # toolsBulletsList = ListFlowable(
        #     toolsBullets, 
        #     bulletType='bullet', 
        #     leftIndent=0.3*inch, 
        #     bulletIndent=0.1*inch, 
        #     spaceAfter=0.1*inch, 
        # )
        
        # hazardsTitle = Paragraph('Hazards:', styles["Title"])
        # hazardsBullets = [ListItem(Paragraph(tool, styles['Normal']), bulletType='bullet', bulletFontSize=styles['Normal'].fontSize) for tool in ptw.hazards]
        # hazardsBulletsList = ListFlowable(
        #     hazardsBullets, 
        #     bulletType='bullet', 
        #     leftIndent=0.3*inch, 
        #     bulletIndent=0.1*inch, 
        #     spaceAfter=0.1*inch, 
        # )

        # controlsTitle = Paragraph('Controls:', styles["Title"])
        # controlsBullets = [ListItem(Paragraph(tool, styles['Normal']), bulletType='bullet', bulletFontSize=styles['Normal'].fontSize) for tool in ptw.controls]
        # controlsBulletsList = ListFlowable(
        #     controlsBullets, 
        #     bulletType='bullet', 
        #     leftIndent=0.3*inch, 
        #     bulletIndent=0.1*inch, 
        #     spaceAfter=0.1*inch, 
        # )

        # risksTitle = Paragraph('Risks:', styles["Title"])
        # risksBullets = [ListItem(Paragraph(tool, styles['Normal']), bulletType='bullet', bulletFontSize=styles['Normal'].fontSize) for tool in ptw.risks]
        # risksBulletsList = ListFlowable(
        #     risksBullets, 
        #     bulletType='bullet', 
        #     leftIndent=0.3*inch, 
        #     bulletIndent=0.1*inch, 
        #     spaceAfter=0.1*inch, 
        # )

        doc.build(elements, onFirstPage=pageHeaderAndWatermark, onLaterPages=pageHeaderAndWatermark, canvasmaker=NumberedCanvas)
        buffer.seek(0)
        main_bytes = buffer.read()
        deiso_bytes = ReportGenerator._buildDeIsolationPdf(ptw)
        if deiso_bytes:
            writer = PdfWriter()
            for src in (main_bytes, deiso_bytes):
                for page in PdfReader(io.BytesIO(src)).pages:
                    writer.add_page(page)
            merged_buf = io.BytesIO()
            writer.write(merged_buf)
            merged_buf.seek(0)
            final_bytes = merged_buf.read()
        else:
            final_bytes = main_bytes
        with tempfile.NamedTemporaryFile(delete=False, prefix=f'ptw-{ptw.id}-', suffix='.pdf') as ptwPdfFile:
            ptwPdfFile.write(final_bytes)
            ptwPdfFile.flush()
            ReportGenerator.openPDF(ptwPdfFile.name)
        # ReportGenerator.MOSReport(ptw.mos, ptw.id, ptw.description)
        if ptw.miwi:
            err, filepath = ClientRequests.getMIWI(loggedUser, ptw.miwi)
            if err:
                return err
            else:
                ReportGenerator.openPDF(filepath)
        ReportGenerator.riskAssessmentReport(ptw.risks, ptw.id, ptw.description)
        
        err, attachs = ClientRequests.getPtwAttachmentNames(loggedUser, ptw.id)
        if err:
            attachs = []
        
        for attachment in attachs:
            err, filepath = ClientRequests.getPtwAttachment(loggedUser, ptw.id, attachment)
            if err:
                print(f"Failed to fetch attachment '{attachment}': {err}")
            else:
                ReportGenerator.openPDF(filepath)
            
        for doc in ptw.requiredDocsToPrint():
            continue
            ReportGenerator.openPDF(doc)

        os.remove(f"qr-summery-{ptw.id}.png")
        return None
    

    def _isolationsTable(ptw, isolations, dataTableWidth, isolate: bool = True):
        styles = getSampleStyleSheet()

        styles['Title'].fontSize = 24
        styles['Title'].leading = 26
        styles["Title"].alignment = TA_LEFT

        styles['Normal'].fontSize = 16
        styles['Normal'].leading = 18
        styles['Normal'].fontName = 'Helvetica'

        styles['Heading3'].fontSize = 16
        styles['Heading3'].leading = 18
        styles['Heading3'].fontName = 'Helvetica-Bold'

        hdr = ParagraphStyle('IsoHdr', parent=styles['Heading3'], fontSize=14, leading=16, alignment=TA_CENTER)
        cel = ParagraphStyle('IsoCell', parent=styles['Normal'],  fontSize=12, leading=14, alignment=TA_CENTER)

        colWeights = [2, 4, 9, 5, 5, 5]
        colWeightsSum  = sum(colWeights)

        ignore_style = ParagraphStyle('IsoIgn', parent=styles['Normal'], fontSize=12, leading=14, alignment=TA_CENTER, textColor=colors.Color(0.6, 0, 0))
        ignore_message = 'Already isolated on PTW:' if isolate else 'Keep isolated for PTW:'

        rows = [[
            Paragraph('No.',         hdr),
            Paragraph('Tag',         hdr),
            Paragraph('Description', hdr),
            Paragraph('Isolator',    hdr),
            Paragraph('Signature',   hdr),
            Paragraph('Date / Time', hdr),
        ]]
        span_cmds = []
        for i, iso in enumerate(isolations, start=1):
            row_idx = len(rows)
            isolated_on = None
            try:
                isolated_on = globalData.isolations[iso.tag].primary_ptw
            except Exception:
                pass

            requiresIsolation = set()
            isReallyActive = False
            try:
                iso_state = globalData.isolations[iso.tag]
                requiresIsolation = set(iso_state.linked_ptws) | set(iso_state.held_by)
                isReallyActive = iso_state.isReallyActive()
            except Exception:
                pass

            ignore = (
                (isolate and (iso.tag in ptw.keep_isolations or (isolated_on and str(ptw.id) != str(isolated_on)))) or
                (not isolate and isReallyActive)
            )

            if ignore:
                rows.append([
                    Paragraph(str(i),                cel),
                    Paragraph(iso.tag or '',         cel),
                    Paragraph(_html.escape(iso.description or ''), cel),
                    Paragraph(f'{ignore_message} {isolated_on if isolate else ", ".join(sorted(requiresIsolation))}', ignore_style),
                    Paragraph('', cel),
                    Paragraph('', cel),
                ])
                span_cmds.append(('SPAN',            (3, row_idx), (5, row_idx)))
                # span_cmds.append(('BACKGROUND',      (3, row_idx), (5, row_idx), colors.Color(1, 0.95, 0.95)))
            else:
                rows.append([
                    Paragraph(str(i),                cel),
                    Paragraph(iso.tag or '',         cel),
                    Paragraph(_html.escape(iso.description or ''), cel),
                    Paragraph(''),
                    Paragraph(''),
                    Paragraph(''),
                ])

        table = Table(rows, repeatRows=1, colWidths=[dataTableWidth * w / colWeightsSum for w in colWeights])
        table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0),  colors.Color(0, 0, 0, 0.15)),
            ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID',          (0, 0), (-1, -1), 0.5, colors.black),
            *span_cmds,
        ]))
        return table


    def openPDF(filepath: str):
        webbrowser.open_new_tab(filepath)
        return

        if platform.system() == "Windows":
            os.startfile(filepath)
        elif platform.system() == "Darwin":  # macOS
            subprocess.call(["open", filepath])
        else:  # Linux
            subprocess.call(["xdg-open", filepath])

        # webbrowser.open_new(filepath)
        # absPath = os.path.abspath(filepath)
        # url = 'file://' + absPath
        # try:
        #     from urllib.request import pathname2url
        # except ImportError:
        #     from urllib import pathname2url
        # url = 'file:{}'.format(pathname2url(absPath))
        # webbrowser.open_new_tab(url)


    def exportPTWs(ptws: list[PTWData]):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'PTWs'

        
        headers = ['PTW#', 'Type', 'Status', 'Date', 'Department', 'Requestor', 'PA', 'Location', 'Area Class', 'Equipment', 'Description']
        col_widths = [8, 20, 15, 14, 18, 22, 22, 20, 12, 25, 50]

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='2E4057')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='AAAAAA')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 28

        def qcolor_to_hex(qcolor):
            r, g, b, a = qcolor.red(), qcolor.green(), qcolor.blue(), 240
            if a < 255:
                r = (r * a + 255 * (255 - a)) // 255
                g = (g * a + 255 * (255 - a)) // 255
                b = (b * a + 255 * (255 - a)) // 255
            return f'{r:02X}{g:02X}{b:02X}'

        cell_align = Alignment(vertical='center', wrap_text=True)

        for row_idx, ptw in enumerate(ptws, start=2):
            status = str(ptw.running_status if ptw.approval_status == PTWData.ApprovalStatus.APPROVED and ptw.running_status is not None else ptw.approval_status)
            requestor = globalData.allUsers[ptw.requestor].getName() if ptw.requestor in globalData.allUsers else str(ptw.requestor or '')
            pa = globalData.allUsers[ptw.performing].getName() if ptw.performing in globalData.allUsers else str(ptw.performing or '')

            row_data = [
                ptw.id,
                str(ptw.type or ''),
                status,
                str(ptw.date or ''),
                str(ptw.department or ''),
                requestor,
                pa,
                str(ptw.location or ''),
                str(ptw.area_class or ''),
                str(ptw.equipment or ''),
                str(ptw.description or ''),
            ]

            row_fill = PatternFill(fill_type='solid', fgColor=qcolor_to_hex(ptw.backgroundColor()))
            row_font = Font(color=qcolor_to_hex(ptw.foregroundColor()))
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_align
                cell.border = border
                cell.fill = row_fill
                cell.font = row_font

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        timestamp = datetime.now().strftime('%d-%m-%Y_%H-%M-%S')
        with tempfile.NamedTemporaryFile(delete=False, prefix=f'ptws-export-{timestamp}-', suffix='.xlsx') as f:
            wb.save(f.name)
            if platform.system() == 'Windows':
                os.startfile(f.name)
            elif platform.system() == 'Darwin':
                subprocess.call(['open', f.name])
            else:
                subprocess.call(['xdg-open', f.name])


    def MOSReport(mos: str, ptwId: str, ptwTitle: str):
        LOGO_IMG_WIDTH = 30*mm
        MARGIN = 0.8 * inch

        if not mos:
            return

        buffer = io.BytesIO()
        timestamp = datetime.now().strftime('%d-%m-%Y - %H:%M:%S')
        doc = SimpleDocTemplate(buffer, pagesize=portrait(A4), leftMargin=MARGIN, rightMargin=MARGIN, topMargin=1.5*MARGIN+LOGO_IMG_WIDTH, bottomMargin=MARGIN)

        pageWidth, pageHeight = portrait(A4)
        headerTableWidth = pageWidth - 2*MARGIN

        styles = getSampleStyleSheet()
        
        styles['Title'].fontSize = 20
        styles['Title'].leading = 20
        styles["Title"].alignment = TA_CENTER

        styles['Normal'].fontSize = 16
        styles['Normal'].leading = 16
        styles['Normal'].fontName = 'Helvetica'

        styles['Heading3'].fontSize = 18
        styles['Heading3'].leading = 18
        styles['Heading3'].fontName = 'Helvetica-Bold'
        styles['Heading3'].alignment = TA_CENTER

        def pageHeaderAndWatermark(canvas: canvas.Canvas, doc):
            canvas.saveState()

            logo1 = Image("./rashpetco-logo.png", LOGO_IMG_WIDTH, LOGO_IMG_WIDTH)
            logo2 = Image("./burullus-logo.png",  LOGO_IMG_WIDTH, LOGO_IMG_WIDTH)
            label = Paragraph(f'MOS for PTW# {ptwId} <br/>' + ptwTitle, styles['Title'])

            table = Table([[logo1, label, logo2]], colWidths=[1.1*LOGO_IMG_WIDTH, headerTableWidth - 2.2*LOGO_IMG_WIDTH, 1.1*LOGO_IMG_WIDTH])
            table.setStyle(TableStyle([
                # Alignment
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), 
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), 

                # grid (optional)
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            table.wrapOn(canvas, headerTableWidth, LOGO_IMG_WIDTH*1.2)
            table.drawOn(canvas, MARGIN, pageHeight - MARGIN - LOGO_IMG_WIDTH)

            canvas.setFont('Helvetica-Bold', 50)
            canvas.setFillColorRGB(0, 0, 0, 0.2)
            canvas.translate(pageWidth / 2.0, pageHeight / 2.0)
            canvas.rotate(55)
            canvas.drawCentredString(0, 0, f'Printed @ {timestamp}')

            canvas.restoreState()
        
        def listToBullets(data: list, style):
            if not data:
                return None
            
            bullets = [ListItem(Paragraph(_html.escape(element), style), bulletType='bullet', bulletFontSize=style.fontSize) for element in data]
            return ListFlowable(
                bullets, 
                bulletType='bullet', 
                leftIndent=0.3*inch, 
                bulletIndent=0.1*inch, 
                spaceAfter=0.1*inch, 
            )
        
        elements = []
        mosSteps = mos.split('\n')
        elements.append(Paragraph('Method of Statement:', styles['Title']))
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(listToBullets(mosSteps, styles['Normal']))

        doc.build(elements, onFirstPage=pageHeaderAndWatermark, onLaterPages=pageHeaderAndWatermark)
        buffer.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, prefix=f'mos-{ptwId}-', suffix='.pdf') as mosPdfFile:
            mosPdfFile.write(buffer.read())
            mosPdfFile.flush()
            ReportGenerator.openPDF(mosPdfFile.name)
        

    def riskAssessmentReport(risksTitles: Iterable[str], ptwId: str, ptwTitle: str):
        LOGO_IMG_WIDTH = 35*mm
        # Cols: No | Hazard | Effect | S | L | Risk(free) | Control | S | L | Risk(ctrl) | Evaluation
        TABLE_WIDTH_WEIGHTS = [7, 30, 33, 4, 4, 9, 52, 5, 5, 10, 20]
        TABLE_WIDTH_WEIGHTS_SUM = sum(TABLE_WIDTH_WEIGHTS)
        MARGIN = 0.35 * inch

        risksTitles = [rt for rt in risksTitles if rt in globalData.allRiskAssessments] 
        if not risksTitles:
            return

        buffer = io.BytesIO()
        timestamp = datetime.now().strftime('%d-%m-%Y - %H:%M:%S')
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=2*MARGIN, rightMargin=2*MARGIN, topMargin=1.5*MARGIN+LOGO_IMG_WIDTH, bottomMargin=MARGIN)

        pageWidth, pageHeight = landscape(A4)
        dataTableWidth = pageWidth - 4*MARGIN

        styles = getSampleStyleSheet()
        
        styles['Title'].fontSize = 20
        styles['Title'].leading = 20
        styles["Title"].alignment = TA_CENTER

        styles['Normal'].fontSize = 12
        styles['Normal'].leading = 12
        styles['Normal'].fontName = 'Helvetica'

        styles['Heading3'].fontSize = 14
        styles['Heading3'].leading = 14
        styles['Heading3'].fontName = 'Helvetica-Bold'
        styles['Heading3'].alignment = TA_CENTER

        styles.add(ParagraphStyle('NormalCenter',  parent=styles['Normal'], alignment=TA_CENTER))
        styles.add(ParagraphStyle('NormalJustify', parent=styles['Normal'], alignment=TA_LEFT))
        _bw = pdfmetrics.stringWidth('• ', 'Helvetica', 12)
        styles.add(ParagraphStyle('BulletItem',    parent=styles['NormalJustify'], leftIndent=_bw, firstLineIndent=-_bw))

        def slr(text):
            t = text or ''
            return t[0], t[-1], t

        def bulleted(text):
            lines = [l for l in (text or '').split('\n') if l.strip()]
            return [Paragraph('• ' + line, styles['BulletItem']) for line in lines] if lines else []

        data = [
            # Row 0 — group headers (non-split cols span both rows via SPAN)
            [
                Paragraph('No.',                  styles['Heading3']), 
                Paragraph('Hazard',               styles['Heading3']), 
                Paragraph('Effect',               styles['Heading3']),
                Paragraph('Free Analysis',        styles['Heading3']), 
                Paragraph('',                     styles['Heading3']), 
                Paragraph('',                     styles['Heading3']),
                Paragraph('Control',              styles['Heading3']),
                Paragraph('Controlled Analysis',  styles['Heading3']), 
                Paragraph('',                     styles['Heading3']), 
                Paragraph('',                     styles['Heading3']),
                Paragraph('Evaluation',           styles['Heading3']), 
            ],
            # Row 1 — sub-headers for split cols
            [
                Paragraph('',                     styles['Heading3']), 
                Paragraph('',                     styles['Heading3']), 
                Paragraph('',                     styles['Heading3']),
                Paragraph('S',                    styles['Heading3']), 
                Paragraph('L',                    styles['Heading3']), 
                Paragraph('Risk',                 styles['Heading3']),
                Paragraph('',                     styles['Heading3']),
                Paragraph('S',                    styles['Heading3']), 
                Paragraph('L',                    styles['Heading3']), 
                Paragraph('Risk',                 styles['Heading3']),
                Paragraph('',                     styles['Heading3']), 
            ],
        ]
        for riskTitle in risksTitles:
            if riskTitle not in globalData.allRiskAssessments:
                continue
            for riskItem in globalData.allRiskAssessments[riskTitle].risks:
                sf, lf, rf = slr(riskItem.free_analysis)
                sc, lc, rc = slr(riskItem.ctrl_analysis)
                data.append([
                    Paragraph(str(len(data) - 1),           styles['NormalCenter']),
                    bulleted(riskItem.hazard),
                    bulleted(riskItem.effect),
                    Paragraph(sf,                           styles['NormalCenter']),
                    Paragraph(lf,                           styles['NormalCenter']),
                    Paragraph(rf,                           styles['NormalCenter']),
                    bulleted(riskItem.ctrl),
                    Paragraph(sc,                           styles['NormalCenter']),
                    Paragraph(lc,                           styles['NormalCenter']),
                    Paragraph(rc,                           styles['NormalCenter']),
                    Paragraph(riskItem.eval,                styles['NormalCenter']),
                ])

        table = Table(data, repeatRows=2, colWidths=[w * dataTableWidth / TABLE_WIDTH_WEIGHTS_SUM for w in TABLE_WIDTH_WEIGHTS])
        table.setStyle(TableStyle([
            # header background (both rows)
            ('BACKGROUND', (0, 0), (-1, 1), colors.Color(0, 0, 0, 0.2)),

            # rowspan for non-split header cols
            ('SPAN', (0,  0), (0,  1)),   # No.
            ('SPAN', (1,  0), (1,  1)),   # Hazard
            ('SPAN', (2,  0), (2,  1)),   # Effect
            ('SPAN', (6,  0), (6,  1)),   # Control
            ('SPAN', (10, 0), (10, 1)),   # Evaluation

            # colspan for group headers
            ('SPAN', (3, 0), (5, 0)),     # Free Analysis
            ('SPAN', (7, 0), (9, 0)),     # Controlled Analysis

            # Alignment
            ('ALIGN',  (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

            # padding
            ('TOPPADDING',    (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING',   (0, 0), (-1, -1), 3),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 3),

            # grid
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements = [table]

        def pageHeaderAndWatermark(canvas: canvas.Canvas, doc):
            canvas.saveState()

            # canvas.drawImage("./rashpetco-logo.png", MARGIN + 2.0 * pageWidth / 3.0, pageHeight - MARGIN - LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, mask='auto')
            # canvas.drawImage("./burullus-logo.png",  MARGIN + 1.0 * pageWidth / 3.0, pageHeight - MARGIN - LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, mask='auto')

            logo1 = Image("./rashpetco-logo.png", LOGO_IMG_WIDTH, LOGO_IMG_WIDTH)
            logo2 = Image("./burullus-logo.png",  LOGO_IMG_WIDTH, LOGO_IMG_WIDTH)
            label = Paragraph(f'RA for PTW# {ptwId} <br/>' + _html.escape(ptwTitle), styles['Title'])

            table = Table([[logo1, label, logo2]], colWidths=[1.2*LOGO_IMG_WIDTH, dataTableWidth - 2.4*LOGO_IMG_WIDTH, 1.2*LOGO_IMG_WIDTH])
            table.setStyle(TableStyle([
                # Alignment
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), 
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), 

                # grid (optional)
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            table.wrapOn(canvas, dataTableWidth, LOGO_IMG_WIDTH*1.2)
            table.drawOn(canvas, 2*MARGIN, pageHeight - MARGIN - LOGO_IMG_WIDTH)

            canvas.setFont('Helvetica-Bold', 50)
            canvas.setFillColorRGB(0, 0, 0, 0.2)
            canvas.translate(pageWidth / 2.0, pageHeight / 2.0)
            canvas.rotate(35)
            canvas.drawCentredString(0, 0, f'Printed @ {timestamp}')

            canvas.restoreState()
        
        doc.build(elements, onFirstPage=pageHeaderAndWatermark, onLaterPages=pageHeaderAndWatermark)
        buffer.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, prefix=f'risk-{ptwId}-', suffix='.pdf') as riskPdfFile:
            riskPdfFile.write(buffer.read())
            riskPdfFile.flush()
            ReportGenerator.openPDF(riskPdfFile.name)
        

    def _buildDeIsolationPdf(ptw: PTWData):
        """Returns PDF bytes for the de-isolation report, or None if not applicable."""
        if not ptw.isolations:
            return None
        if ptw.running_status not in (PTWData.RunningStatus.HELD, PTWData.RunningStatus.CLOSED):
            return None

        QR_CODE_WIDTH = 60*mm
        LOGO_IMG_WIDTH = 40*mm
        MARGIN = 0.6 * inch

        buffer = io.BytesIO()
        timestamp = datetime.now().strftime('%d-%m-%Y - %H:%M:%S')

        pageWidth, pageHeight = landscape(A4)
        dataTableWidth = pageWidth - 2*MARGIN - QR_CODE_WIDTH

        styles = getSampleStyleSheet()
        styles['Title'].fontSize = 24
        styles['Title'].leading = 26
        styles['Title'].alignment = TA_LEFT
        styles['Normal'].fontSize = 16
        styles['Normal'].leading = 18
        styles['Normal'].fontName = 'Helvetica'
        styles['Heading3'].fontSize = 16
        styles['Heading3'].leading = 18
        styles['Heading3'].fontName = 'Helvetica-Bold'

        ReportGenerator._makeQrWithLogo(ptw)

        isolationsByType = {}
        for iso in ptw.isolations:
            if iso.type not in isolationsByType:
                isolationsByType[iso.type] = []
            isolationsByType[iso.type].append(iso)

        try:
            pdfmetrics.registerFont(TTFont('Satisfy', './fonts/Satisfy/Satisfy-Regular.ttf'))
            sig_font = 'Satisfy'
        except Exception:
            sig_font = 'Helvetica-Oblique'

        if ptw.running_status == PTWData.RunningStatus.HELD:
            pa_user = globalData.allUsers[ptw.hold_performing].getName() if ptw.hold_performing in globalData.allUsers else str(ptw.hold_performing or '')
            ia_user = globalData.allUsers[ptw.hold_issuing].getName()    if ptw.hold_issuing    in globalData.allUsers else str(ptw.hold_issuing    or '')
            sig_cols = [
                ('PA', pa_user, ptw.hold_performing_timestamp or ''),
                ('IA', ia_user, ptw.hold_issuing_timestamp    or ''),
            ]
        else:
            pa_user = globalData.allUsers[ptw.close_performing].getName() if ptw.close_performing in globalData.allUsers else str(ptw.close_performing or '')
            ia_user = globalData.allUsers[ptw.close_issuing].getName()    if ptw.close_issuing    in globalData.allUsers else str(ptw.close_issuing    or '')
            sig_cols = [
                ('PA', pa_user, ptw.close_performing_timestamp or ''),
                ('IA', ia_user, ptw.close_issuing_timestamp    or ''),
            ]

        SIG_TRAILER_HEIGHT = 2.0 * inch
        left_x = MARGIN + QR_CODE_WIDTH

        label_style = ParagraphStyle('SigLabel', parent=styles['Heading3'], fontSize=16, leading=18, alignment=TA_CENTER)
        sig_style   = ParagraphStyle('Signature', parent=styles['Normal'],  fontSize=16, leading=18, alignment=TA_CENTER, fontName=sig_font)
        date_style  = ParagraphStyle('SigDate',   parent=styles['Normal'],  fontSize=14, leading=16, alignment=TA_CENTER)

        def sigTables(columns: list, chunk_size=3):
            chunks = [columns[i:i+chunk_size] for i in range(0, len(columns), chunk_size)]
            result = []
            for chunk in chunks:
                col_w = dataTableWidth / len(chunk)
                header_row, name_row, date_row = [], [], []
                style_cmds = [
                    ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING',    (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]
                for i, col in enumerate(chunk):
                    if col is not None:
                        label, sig, ts = col
                        style_cmds.append(('LINEBELOW', (i, 0), (i, 0), 0.5, colors.black))
                        style_cmds.append(('LINEBELOW', (i, 1), (i, 1), 0.5, colors.black))
                    else:
                        label, sig, ts = '', '', ''
                    header_row.append(Paragraph(label or ' ', label_style))
                    name_row.append(  Paragraph(sig   or ' ', sig_style))
                    date_row.append(  Paragraph(ts    or ' ', date_style))
                table = Table(
                    [header_row, name_row, date_row],
                    colWidths=[col_w] * len(chunk),
                )
                table.setStyle(TableStyle(style_cmds))
                result.append(table)
                result.append(Spacer(1, 0.2 * inch))
            return result

        def pageHeaderAndWatermark(canv, doc):
            canv.saveState()
            canv.drawImage('./rashpetco-logo.png', 0.7 * MARGIN + (QR_CODE_WIDTH - LOGO_IMG_WIDTH) / 2, (pageHeight - LOGO_IMG_WIDTH) / 2.0 + pageHeight / 3.5, LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, mask='auto')
            canv.drawImage('./burullus-logo.png',  0.7 * MARGIN + (QR_CODE_WIDTH - LOGO_IMG_WIDTH) / 2, (pageHeight - LOGO_IMG_WIDTH) / 2.0 - pageHeight / 3.5, LOGO_IMG_WIDTH, LOGO_IMG_WIDTH, mask='auto')
            canv.setFont('Helvetica-Bold', 50)
            canv.setFillColorRGB(0, 0, 0, 0.2)
            canv.translate(pageWidth / 2.0, pageHeight / 2.0)
            canv.rotate(35)
            canv.drawCentredString(0, 0, f'Printed @ {timestamp}')
            canv.restoreState()
            canv.drawImage(f'qr-summery-{ptw.id}.png', 0.7 * MARGIN, (pageHeight - QR_CODE_WIDTH) / 2.0, QR_CODE_WIDTH, QR_CODE_WIDTH)

            sig_flowables = [
                Paragraph('De-Isolation Confirmations:', styles['Title']),
                Spacer(1, 0.07 * inch),
                *sigTables(sig_cols),
            ]
            sig_frame = Frame(
                left_x, MARGIN * 0.7,
                dataTableWidth, SIG_TRAILER_HEIGHT,
                leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
            )
            sig_frame.addFromList(sig_flowables, canv)

        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            leftMargin=MARGIN + QR_CODE_WIDTH, rightMargin=MARGIN,
            topMargin=MARGIN * 0.7, bottomMargin=MARGIN * 0.7 + SIG_TRAILER_HEIGHT,
        )

        elements = []
        for isoType in isolationsByType:
            elements.extend([
                Paragraph(f'{isoType} De-Isolations:', styles['Title']),
                Spacer(1, 0.07 * inch),
                ReportGenerator._isolationsTable(ptw, isolationsByType[isoType], dataTableWidth, isolate=False),
                PageBreak(),
            ])

        doc.build(elements, onFirstPage=pageHeaderAndWatermark, onLaterPages=pageHeaderAndWatermark)
        buffer.seek(0)
        return buffer.read()


    def deIsolationReport(loggedUser, ptw: PTWData):
        pdf_bytes = ReportGenerator._buildDeIsolationPdf(ptw)
        if pdf_bytes is None:
            return
        with tempfile.NamedTemporaryFile(delete=False, prefix=f'ptw-{ptw.id}-', suffix='.pdf') as f:
            f.write(pdf_bytes)
            f.flush()
            ReportGenerator.openPDF(f.name)

