import secrets
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem,
                              QAbstractItemView, QHeaderView, QDialogButtonBox)
from User import User, UserRoles, UserDepartments
from utils import parseTabularFile


class DialogUsersPreview(QDialog):
    def __init__(self, parent, title: str, headers: list[str], rows: list[list],
                 mode: str = 'confirm', summary: str = '', onExport=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(750, 400)

        lyt = QVBoxLayout()
        self.setLayout(lyt)

        if summary:
            lblSummary = QLabel(summary)
            lblSummary.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lyt.addWidget(lblSummary)

        self.tbl = QTableWidget()
        self.tbl.setColumnCount(len(headers))
        self.tbl.setHorizontalHeaderLabels(headers)
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.tbl.horizontalHeader().setStretchLastSection(True)
        self.tbl.verticalHeader().hide()

        self.tbl.setRowCount(len(rows))
        for r, record in enumerate(rows):
            for c, value in enumerate(record):
                self.tbl.setItem(r, c, QTableWidgetItem(str(value)))
        lyt.addWidget(self.tbl)

        btns = QDialogButtonBox()
        if mode == 'confirm':
            btnOk = btns.addButton("Import", QDialogButtonBox.ButtonRole.AcceptRole)
            btnCancel = btns.addButton(QDialogButtonBox.StandardButton.Cancel)
            btnOk.clicked.connect(self.accept)
            btnCancel.clicked.connect(self.reject)
        else:
            if onExport:
                btnExport = btns.addButton("Export as Excel", QDialogButtonBox.ButtonRole.ActionRole)
                btnExport.clicked.connect(onExport)
            btnClose = btns.addButton(QDialogButtonBox.StandardButton.Close)
            btnClose.clicked.connect(self.accept)
        lyt.addWidget(btns)


HEADERS = ['Username', 'Name', 'Role', 'Department', 'Email', 'EXT']
FIELDS = ['username', 'name', 'role', 'department', 'email', 'ext']


class ImportRow:
    def __init__(self, rowNum: int, username: str, name: str, role: str, department: str,
                 email: str, ext: str, user: User = None, status: str = ''):
        self.rowNum = rowNum
        self.username = username
        self.name = name
        self.role = role
        self.department = department
        self.email = email
        self.ext = ext
        self.user = user
        self.status = status

    def asRecord(self) -> list:
        return [
            self.username, self.name, self.role, self.department, self.email, self.ext,
            self.user.getPassword() if self.user else '',
            self.status,
        ]


class ImportUsersExcel:
    HEADERS = HEADERS
    FIELDS = FIELDS

    @staticmethod
    def parseFile(filepath: str, existingUsernames: set) -> list[ImportRow]:
        dataRows = parseTabularFile(filepath, HEADERS)

        roleValues = {r.value.lower(): r.value for r in UserRoles}
        deptValues = {d.value.lower(): d.value for d in UserDepartments}

        rows = []
        seenUsernames = set()

        for rowNum, record in enumerate(dataRows, start=2):
            if all(not v for v in record):
                continue

            data = dict(zip(FIELDS, record))

            username = data['username']
            name = data['name']
            role = data['role']
            department = data['department']

            error = None
            if not username:
                error = "Username can't be empty"
            elif username in existingUsernames or username in seenUsernames:
                error = f"Username '{username}' already exists"
            elif not name:
                error = "Name can't be empty"
            elif role.lower() not in roleValues:
                error = f"Invalid role '{role}'"
            elif department.lower() not in deptValues:
                error = f"Invalid department '{department}'"

            user = None
            if error is None:
                seenUsernames.add(username)
                user = User(
                    username=username,
                    password=secrets.token_urlsafe(12),
                    name=name,
                    role=roleValues[role.lower()],
                    department=deptValues[department.lower()],
                    email=data['email'],
                )
                user.setExt(data['ext'])

            rows.append(ImportRow(
                rowNum=rowNum,
                username=username,
                name=name,
                role=role,
                department=department,
                email=data['email'],
                ext=data['ext'],
                user=user,
                status="Ready to import" if user else f"Skipped: {error}",
            ))

        return rows

    @staticmethod
    def exportResult(filepath: str, rows: list[ImportRow]):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Import Result'

        headers = HEADERS + ['Password', 'Status']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        for rowIdx, row in enumerate(rows, start=2):
            for col, value in enumerate(row.asRecord(), start=1):
                ws.cell(row=rowIdx, column=col, value=value)

        wb.save(filepath)
